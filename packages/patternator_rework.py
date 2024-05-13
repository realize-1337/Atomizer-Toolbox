import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.figure_factory as ff
from scipy.optimize import curve_fit
from scipy.integrate import quad
from sklearn.metrics import r2_score
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.comments import Comment
import os

def createExcel(path, cols=114, tests=50, steps=3, first=['s_l', 'm_g,i', 'm_l',' m_g,o', 'rho_l', 'time [s]', 'PDA']):
    numbers = [x+1 for x in range(cols)]
    header = [*first, *numbers]
    tooltips = ['Liquid sheet thickness or nozzle name', 'Inner gas mass flow rate [kg h-1]', 'Liquid mass flow rate (important!) [kg h-1]', 
                'Outer gas mass flow rate [kg h-1]', 'Liquid mass density (important!) [kg m-3]', 'Measuring time [s]', 'Width of PDA measurement (optional) [mm]']
    print(header)

    def draw_thick_horizontal_line(sheet, row, start_col, end_col):
        for col in range(start_col, end_col + 1):
            # for r in range(tests*steps):
            #     sheet.cell(row=r+1, column=len(first)).border = Border(right=Side(style='thick'))
            for r in range(tests+1):
                sheet.cell(row=r*steps+row, column=col).border = Border(bottom=Side(style='thick'))
                if col == len(first):
                    currentRow = r*steps+row
                    sheet.cell(row=r*steps+row, column=col).border = Border(bottom=Side(style='thick'), right=Side(style='thick'))
                    if r != tests:
                        for x in range(1, steps):
                            sheet.cell(row=currentRow+x, column=col).border = Border(right=Side(style='thick'))    
    def set_row_values(sheet, row_num=1, values:list=[]):
        for col_num, value in enumerate(values, start=1):
            sheet.cell(row=row_num, column=col_num).value = value
            sheet.cell(row=row_num, column=col_num).font = Font(bold=True)

    def setTooltips(sheet, tt, row=1, col=1):
        for i, t in enumerate(tt):
            sheet.cell(row, col+i).comment = Comment(t, 'DM')

    wb = Workbook()
    ws = wb.active

    set_row_values(ws, 1, header)
    draw_thick_horizontal_line(ws, row=1, start_col=1, end_col=len(header))
    setTooltips(ws, tooltips)
    
    wb.save(path)


class pat():
    def __init__(self, path, a=5.5, b=5.5, h=0.35, steps=3) -> None:
        self.path = path
        self.a = a
        self.b = b
        self.h = h
        self.s = steps
        self.current_it = 0
        self.globalread()

    def globalread(self)->None:
        self.gdf = pd.read_excel(self.path, decimal=',')
        self.exportDF = self.gdf.copy().iloc[:,0:7]
        self.exportDF['Width[mm]'] = np.nan
        self.exportDF['exDelta'] = np.nan
        self.exportDF['exDeltaAbs'] = np.nan
        self.exportDF['fitDelta'] = np.nan
        self.exportDF['fitDeltaInPDA'] = np.nan
        self.exportDF['Fit-R2'] = np.nan
        self.it = len(self.gdf)/self.s
        self.gdf = self.gdf.fillna(0)
        self.gdf = self.gdf.applymap(lambda x: 0 if isinstance(x, str) and x.strip() == '' else x)
        arr = self.gdf.to_numpy()[:,7:]
        self.gcols = len(arr[0, :])
        self.garr = arr#.astype(np.float32)

    def read(self)-> None:
        self.df = self.gdf.iloc[self.current_it:self.current_it+self.s]
        # print(self.df)
        print(25*'--')
        print(f"{self.df['s_l'][self.current_it]} - {self.df['m_g,i'][self.current_it]} - {self.df['m_l'][self.current_it]} - {self.df['m_g,o'][self.current_it]} - {self.df['rho_l'][self.current_it]}")
        self.current_it += self.s
        arr = self.df.to_numpy()[:,7:]
        self.cols = len(arr[0, :])
        self.arr = arr.astype(np.float32)
        
    def findCenter(self)->list:
        max = np.argmax(self.arr, axis=1)
        max_2 = []
        for i, m in enumerate(max):
            try:
                off = (np.argmax(self.arr[i, m-1:m+2:1]))
                if off == 1: 
                    max_2.append(m+1)
                else: max_2.append(m-1)
            except:
                max_2.append(1)
        center = []
        for m1, m2 in zip(max, max_2):
            center.append(0.5*(m1+m2))
        return center
    
    def off(self)->list:
        tup = []
        for i in range(len(self.arr)):
            max = np.argmax(self.arr[i, :])
            # print(max)
            left = self.arr[i, max-1]
            right = self.arr[i, max+1]
            delta_left = (self.arr[i, max]-left)/self.arr[i, max]
            delta_right = (self.arr[i, max]-right)/self.arr[i, max]
            tup.append((delta_left, delta_right))

        centers = []
        for i, t in enumerate(tup):
            max = np.argmax(self.arr[i, :])
            l, r = t
            if r == 0: centers.append([2, max, max+1])
            elif l == 0: centers.append([2, max, max-1])
            elif l/r > 2: centers.append([2, max, max-1])
            elif l/r <0.5: centers.append([2, max, max+1])
            else: centers.append([1, max, None])
        print(centers)
        return centers

    def calcNewArea(self, mode=2):
        delta = self.a+self.h
        if mode == 2:
            colCenters = np.arange(0.5*self.h+0.5*self.a, self.cols*delta, delta)
            areas = np.zeros_like(colCenters)
            areas = np.pi*((colCenters+0.5*self.a)**2-(colCenters-0.5*self.a)**2)/2
        else:
            colCenters = np.arange(0, self.cols*delta, delta)
            areas = np.zeros_like(colCenters)
            areas = np.pi*((colCenters+0.5*self.a)**2-(colCenters-0.5*self.a)**2)/2
            areas[0] = np.pi*(0.5*self.a)**2/2
        return areas
     
    def exData(self, glue=2, printIndividual=False, diagrams=True, diagrams_height=False):
        centers = self.off()
        liq = self.df['m_l'].to_numpy()
        time = self.df['time [s]'].to_numpy()
        dens = self.df['rho_l'].to_numpy()
        masses = []
        fluxs = []
        rel = []
        zero = int(self.current_it-self.s)
        PDA_w = self.df['PDA'].to_numpy()
        inPDAList = []
        lossRel = []
        lossAbs = []
        for i, trip in enumerate(centers):
            m, c, c2 = trip
            h = self.arr[i, :]
            if dens[i] == 1236.27:
                h = h*(1-0.0431)
            if dens[i] == 1222.96:
                h = h*(1-0.0163)
            
            flux = h*1e-9*dens[i]/time[i]

            if m ==2:
                if c < c2: # Max ist rechts
                        left = flux[:c][::-1]
                        right = flux[c:]
                else: # Max is links
                    left = flux[:c+1][::-1]
                    right = flux[c+1:]

                areas = self.calcNewArea(m)

            else: 
                left = np.copy(flux[:c+1][::-1])
                right = np.copy(flux[c:])

                areas = self.calcNewArea(m)
            
            

            massL = left*areas[:len(left)]
            massR = right*areas[:len(right)]
            total = (np.sum(massL) + np.sum(massR))*3600
            mass = np.concatenate((massL[::-1], massR))
            width = int(np.round(0.5*PDA_w[i]/(self.a+self.h)))
            inPDA = (np.sum(massL[:width]) + np.sum(massR[:width]))*3600
            inPDA_rel = inPDA/total
            inPDAList.append(inPDA_rel)


            input = liq[i]
            diff = input-total
            lossAbs.append(diff)
            if printIndividual:
                print(f'Relative Difference: {"%.3f" % (diff/input*100)}%')
            self.exportDF['exDelta'].iloc[zero+i] = diff/input*100
            self.exportDF['exDeltaAbs'].iloc[zero+i] = diff
            if diff/input*100 < 100:
                rel.append(diff/input*100)
            fluxs.append(flux)
            masses.append(mass)

        print('----')
        print(f'Average: {"%.3f" % np.mean(np.array(rel))}%')
        

        if diagrams:
            t = np.mean(time) 
            data = []
            data_flux = []
            for i, (mass, center, flux) in enumerate(zip(masses, centers, fluxs)):
                mode, c, c2 = center
                delta = self.a+self.h
                if mode == 2: colCenters = np.arange(0.5*self.h+0.5*self.a, self.cols*delta, delta) - 0.5*self.cols*delta
                else: colCenters = np.arange(0, self.cols*delta, delta) - 0.5*self.cols*delta
                data.append(go.Scatter(mode='lines', x=colCenters, y=mass, line_shape='hvh', name=f'Run {i}: Center Points: {m}'))
                data_flux.append(go.Scatter(mode='lines', x=colCenters, y=flux, line_shape='hvh', name=f'Run {i}: Center Points: {m}'))
                
            fig = go.Figure(data)
            fig.write_html(os.path.join(os.path.dirname(self.path), 'ex', f"Ex_{self.df['s_l'][self.current_it-self.s]}_{self.df['m_g,i'][self.current_it-self.s]}_{self.df['m_l'][self.current_it-self.s]}_{self.df['m_g,o'][self.current_it-self.s]}_{self.df['rho_l'][self.current_it-self.s]}.html"))
            fig = go.Figure(data_flux)
            fig.write_html(os.path.join(os.path.dirname(self.path), 'ex_flux', f"Ex_{self.df['s_l'][self.current_it-self.s]}_{self.df['m_g,i'][self.current_it-self.s]}_{self.df['m_l'][self.current_it-self.s]}_{self.df['m_g,o'][self.current_it-self.s]}_{self.df['rho_l'][self.current_it-self.s]}.html"))
            pass

        if diagrams_height:
            t = np.mean(time) 
            data = []
            for i, (h, center) in enumerate(zip(hs, centers)):
                mode, c, c2 = center
                delta = self.a+self.h
                if mode == 2: colCenters = np.arange(0.5*self.h+0.5*self.a, self.cols*delta, delta) - 0.5*self.cols*delta
                else: colCenters = np.arange(0, self.cols*delta, delta) - 0.5*self.cols*delta
                data.append(go.Scatter(mode='lines', x=colCenters, y=h/time[i]*60, line_shape='hvh', name=f'Run {i}: Center Points: {m}'))
                
            fig = go.Figure(data)
            fig.write_html(os.path.join(os.path.dirname(self.path), 'ex_height', f"Ex_{self.df['s_l'][self.current_it-self.s]}_{self.df['m_g,i'][self.current_it-self.s]}_{self.df['m_l'][self.current_it-self.s]}_{self.df['m_g,o'][self.current_it-self.s]}_{self.df['rho_l'][self.current_it-self.s]}.html"))
            pass
        
        return [fluxs, centers]
    
    def fit_lorentz(self, fluxs, centers, r_limit=0.95, res = 1000):
        def find_nearest(array, value):
            array = np.asarray(array)
            idx = (np.abs(array - value)).argmin()
            return idx

        params = []
        integ = []
        r_2 = []
        delta = self.a+self.h
        x_data_pre = np.arange(self.h+0.5*self.a, self.cols*delta, delta)
        time = self.df['time [s]'].to_numpy()
        PDA_w = np.mean(self.df['PDA'].to_numpy())
        print(f'PDA WIDTH: {PDA_w}')
        w = 0
        liq = self.df['m_l'].to_numpy()
        point = int(self.current_it-self.s)
        non_zero_rows = np.any(self.arr != 0, axis=1)

        for i in range(len(fluxs)):
            flux = fluxs[i]
            m, c, c2 = centers[i]

            if m ==2:
                if c < c2: # Max ist rechts
                        left = flux[:c][::-1]
                        right = flux[c:]
                else: # Max is links
                    left = flux[:c+1][::-1]
                    right = flux[c+1:]

            else: 
                left = np.copy(flux[:c+1][::-1])
                right = np.copy(flux[c:])

            
            max = np.max(flux)
            max_point = np.argmax(flux)
            FWHM = abs(find_nearest(left, max/2) + find_nearest(right, max/2)-1)*delta

            popt, pcov, id, mesg, ier  = curve_fit(self.lorentz, xdata=x_data_pre, ydata=flux, method='lm', p0=[max, x_data_pre[max_point], FWHM], xtol=1e-9, gtol=1e-9, full_output=True)
            
            peak, zero, width = popt
            fullWidth = (np.argmin(left) + np.argmin(right)-1)*delta #+ self.a
            x_data = np.linspace(0, 0.5*fullWidth, res)
            x_data = np.arange(0, 0.5*fullWidth+delta, delta)
            diffI = np.diff(x_data, prepend=0)
            last = x_data[-1] + x_data[-1]/(res-1)*0.5
            last = x_data[-1] + 0.5*delta
            diffO = np.diff(x_data, append=last)
            Raw_Area = np.zeros_like(x_data)
            Raw_Area = np.pi * ((x_data+diffO)**2 - (x_data-diffI)**2)
            y_fit = self.lorentz(x_data_pre, *popt)
            r2 = r2_score(flux, y_fit)
            # print(f'R^2 Score: {"%.4f" % r2}')
            y_fit = self.lorentz(x_data, peak, 0, width)
            mass = np.sum(y_fit*Raw_Area)*3600/2
            # print(f'Fitted Mass: {mass}')
            self.exportDF['fitDelta'].iloc[point+i] = (liq[i]-mass)/liq[i]*100
            self.exportDF['Fit-R2'].iloc[point+i] = r2
            self.exportDF['Width[mm]'].iloc[point+i] = fullWidth

            try:
                tMass = (y_fit*Raw_Area)/2
                tMass = np.cumsum(tMass[::-1])
                tMass = tMass/np.max(tMass)
                peak_ = int(np.round(0.5*PDA_w/x_data[1]))
                if peak_ > len(tMass): peak_ = len(tMass)-1
                print(f'PDA Mass Fraction: {"%.2f" % (tMass[peak_]*100)}')
                params.append(popt)
            except: print('FAULTY DATA')

            ### Fitted mass in PDA area to total mass
            
            x_data = np.linspace(0, 0.5*PDA_w, res)
            diffI = np.diff(x_data, prepend=0)
            last = x_data[-1] + x_data[-1]/(res-1)*0.5
            diffO = np.diff(x_data, append=last)
            Raw_Area = np.zeros_like(x_data)
            Raw_Area = np.pi * ((x_data+diffO)**2 - (x_data-diffI)**2)
            # print(f'R^2 Score: {"%.4f" % r2}')
            y_fit = self.lorentz(x_data, peak, 0, width)
            mass = np.sum(y_fit*Raw_Area)*3600/2
            # print(f'Fitted Mass: {mass}')
            self.exportDF['fitDeltaInPDA'].iloc[point+i] = (liq[i]-mass)/liq[i]*100
            
            # self.exportDF['Fit-R2'].iloc[point+i] = r2

        return [params, integ, r_2, w]

    @staticmethod
    def lorentz(x, A, xc, w):
            return A*(w/(4*(x-xc)**2+w**2))
        
    def run(self)->pd.DataFrame:
        for i in range(int(self.it)):
            try:
                self.read()
                fluxs, centers = self.exData(printIndividual=True)

                params, integ, r_2, width = self.fit_lorentz(fluxs, centers)
                params = np.array(params)
                print(np.mean(params, axis=0))
             
            except: pass
        
        self.exportDF = self.exportDF.fillna('n/a')
        return self.exportDF
    
        for i in range(int(self.it)):
            off = self.off()
            for l, r in off:
                ratio = l/r
                if ratio > 1 : ratio = 1/ratio
                print(f'Links/Max: {"%.2f"%(l*100)} % -- Recht/Max: {"%.2f"%(r*100)}  % -- Links/Rechts: {"%.2f"%(ratio)}')
        return None        

if __name__ == '__main__':
    createExcel(r'C:\Users\david\Desktop\excel.xlsx')
    pat = pat(r'C:\Users\david\Documents\Dev\Atomizer-Toolbox\test\patternator\patternator.xlsx')
    df = pat.run()

    # df = pat.run()
    # df.to_excel(r'C:\Users\david\Documents\Dev\Atomizer-Toolbox\test\patternator\patternator_ex_new.xlsx')
