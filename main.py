# All imports for the main.py
import av
import os
import sys
import math
from time import time
from datetime import datetime
import subprocess
import json
import ctypes
import numpy as np
from PIL import Image
import pandas as pd
import pyperclip as pc
from functools import partial
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from PyQt6.QtWidgets import *
from PyQt6 import QtGui
from PyQt6.QtCore import QRunnable, QThreadPool, pyqtSignal, QObject, QTimer, Qt
from PyQt6.QtGui import QPixmap, QPen, QColor
from openpyxl import load_workbook
from openpyxl.comments import Comment
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import packages.dimLess as dL
from packages.calculator import Calculator as ca
import packages.multiAngle as mA
import packages.patternator_rework as pat
from packages.createMatlabScripts import MLS
from pyfluids import Fluid, FluidsList, Input
import logging
UI_FILE = os.path.abspath('GUI\mainWindow.ui')
UI_FILE = 'GUI\mainWindow.ui'
PY_FILE = 'GUI\mainWindow.py'
# subprocess.run(['pyuic6', '-x', UI_FILE, '-o', PY_FILE])
# The subprocess is used to compile the .ui file to a .py file. 
# This is only neccessary after changes in the .ui file with qt designer or qt creator, make sure to comment it out before compiling
from GUI.mainWindow import Ui_MainWindow as main
import packages.exportTable as ex
import packages.bulkExport as bulkex
import packages.PDA as PDA
from packages.settings import settings
from packages.exportDB import exportDB
from skimage import morphology
import cv2
import plotly.graph_objs as go
import plotly.express as px
import logging
import webbrowser
try:
    import pyi_splash
    # pyi_splash.update_text('UI Loaded ...')
    pyi_splash.close()
except:
    pass

# Important:
# Multithreading does NOT work in pyqt. 
# If you want to run processes at the same time, you have to create a worker of class QRunnable
# The corrospondig signal class is requiered for communication between the main process and the multiple processes from the QRunnable class

# Make sure to check the classes below to understand how it works
# I technically created packages for most classes in the package folder, 
# however, this sometimes lead to performace issues during class creation, 
# as a workarround, the packages have been imported manually to some classes but not for all

class WorkerSignals(QObject):
    finished = pyqtSignal() 

class Worker(QRunnable):

    def __init__(self, items:tuple, path, filetype):
        super().__init__()
        self.index, self.frame = items
        self.path = path
        self.filetype = filetype
        self.signals = WorkerSignals()

    def run(self):
        frame = self.frame
        frame = frame.reformat(format='gray')
        img = frame.to_image()

        # pil_image.save(os.path.join(self.path, 'global', 'currentCine', f'frame_{index}.jpg'))
        img.save(os.path.join(self.path, 'global', 'currentCine', f'frame_{"%04d" % self.index}.{self.filetype}'))
        self.signals.finished.emit()

class FreqSignals(QObject):
    push = pyqtSignal(tuple)
    finished = pyqtSignal() 

class FreqWorker(QRunnable):

    def __init__(self, index, file, ref, x_start:int, x_end:int, y:int, refImage=None):
        super().__init__()
        self.path = file
        self.id = index
        self.signals = FreqSignals()
        self.x_start = x_start
        self.x_end = x_end
        self.y = y
        self.ref = refImage

    def correction(self, image):
        mw = np.mean(self.ref)
        pic_cor = np.uint8((np.double(image) / np.double(self.ref)) * mw)
        return pic_cor

    def run(self):
        image = cv2.imread(os.path.normpath(rf'{self.path}'), cv2.IMREAD_GRAYSCALE)
        pic_cor = self.correction(image)
        _, binary_image = cv2.threshold(pic_cor, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)

        minPix = np.count_nonzero(binary_image == 0)/2
        binary_edit = morphology.remove_small_objects(binary_image, min_size=round(minPix))
        
        inverted_image = cv2.bitwise_not(binary_edit)
        contours, _ = cv2.findContours(inverted_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        for contour in contours:
            cv2.drawContours(binary_edit, [contour], 0, 0, -1)

        line = binary_edit[self.y, self.x_start:self.x_end]
        tup = (self.id, np.sum(line == 0))
        self.signals.push.emit(tup)
        self.signals.finished.emit()

class WorkerSignalsConversion(QObject):
    finishedConversion = pyqtSignal() 

class WorkerConversion(QRunnable):
    def __init__(self, file, filetype, keep=True, compression=False, overwrite=True):
        super().__init__()
        self.file = file
        self.keep = keep
        self.overwrite = overwrite
        self.compression = compression
        self.filetype = filetype
        self.signals = WorkerSignalsConversion()

    def run(self):
        container = av.open(self.file)
        path = os.path.dirname(self.file)
        logging.info('Working on it.')
        for index, frame in enumerate(container.decode(video=0)):
            if not self.overwrite:
                if os.path.exists(os.path.join(path, f'frame_{"%04d" % index}.{self.filetype}')): 
                    self.signals.finishedConversion.emit()
                    continue
            frame = frame.reformat(format='gray')
            img = frame.to_image()
            if self.compression:
                img.save(os.path.join(path, f'frame_{"%04d" % index}.{self.filetype}'), compression="jpeg")
            else: img.save(os.path.join(path, f'frame_{"%04d" % index}.{self.filetype}'))
            self.signals.finishedConversion.emit()
        container.close()
        if not self.keep:
            os.remove(self.file)

class DropletsSignals(QObject):
    push = pyqtSignal(tuple)
    diasPush = pyqtSignal(list)
    finished = pyqtSignal() 

class DropletsWorker(QRunnable):

    def __init__(self, path, refImage_gray, threshold:int=40, circ:float=0.6, scale:float = 1):
        super().__init__()
        self.signals = DropletsSignals()
        self.path = path
        self.refImage = refImage_gray
        self.threshold = threshold
        self.circ = circ
        self.scale = scale

    def getImage(self):
        droplets_img = cv2.imread(os.path.normpath(rf'{self.path}'))
        self.droplets_gray = cv2.cvtColor(droplets_img, cv2.COLOR_BGR2GRAY)
        diff_img = cv2.absdiff(self.droplets_gray, self.refImage)
        return diff_img
    
    def getContours(self, diff_img):
        _, self.thresholded_img = cv2.threshold(diff_img, self.threshold, 255, cv2.THRESH_BINARY)
        # contours, _ = cv2.findContours(self.thresholded_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours, _ = cv2.findContours(self.thresholded_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

        round_contours = []
        circularity_threshold = self.circ
        for contour in contours:
            perimeter = cv2.arcLength(contour, True)
            area = cv2.contourArea(contour)
            if perimeter > 0:
                circularity = (4 * 3.1416 * area) / (perimeter * perimeter)
                if circularity > circularity_threshold:
                    round_contours.append(contour)

        return round_contours

    def findLargestRound(self, round_contours):
        largest_area = 0
        for contour in round_contours:
            area = cv2.contourArea(contour)
            if area > largest_area:
                largest_area = area
                largest_contour = contour

        try:
            (x, y), radius = cv2.minEnclosingCircle(largest_contour)
        except:
            x = 0
            y = 0 
            radius = 0

        return x, y, radius
    
    def listAllAbove(self, round_contours):
        dias = []
        for contour in round_contours:
            try: 
                (x, y), radius = cv2.minEnclosingCircle(contour)
                dia = radius*2/self.scale
                if dia > 1452.2/1000:
                    dias.append(dia)
            except: pass
        return dias

    def generateImage(self, x:int, y:int, radius:int):
        circle_img = np.zeros_like(self.droplets_gray)
        if radius < 6:
            radius *= 6
        cv2.circle(circle_img, (int(x), int(y)), int(radius), (255, 255, 255), 2)
        result_img = cv2.cvtColor(self.droplets_gray, cv2.COLOR_GRAY2BGR)
        result_img = cv2.addWeighted(result_img, 1, cv2.cvtColor(circle_img, cv2.COLOR_GRAY2BGR), 0.5, 0)
        return result_img

    def generateReport(self, items:list, export):
        '''
        Items must be a list of paths in the order of diameter size.
        The length is not relevant, however, i might come with a significant performace drawback.
        '''
        threshholds = []
        results = []
        diameters = []
        refSize = self.scale
        for item in items:
            self.path = item
            diff_img = self.getImage()
            round_contours = self.getContours(diff_img)
            x, y, radius = self.findLargestRound(round_contours)
            diameters.append(2*radius/refSize*10**3)
            threshholds.append(self.thresholded_img)
            results.append(self.generateImage(x, y, radius+0.15*radius))

        
        if refSize != 1: unit = 'μm'
        else: unit = 'px'

        figs = []
        for i in range(len(diameters)):
            fileName = items[i].split("\\")[1]
            fig, axs = plt.subplots(2, 1, figsize=(12, 12))
            axs[0].imshow(cv2.cvtColor(results[i], cv2.COLOR_BGR2RGB))
            axs[0].set_title(f'#{i+1} \t {"%.2f" % diameters[i]} {unit} \t {fileName}')
            axs[1].imshow(threshholds[i], cmap='gray')
            axs[1].set_title(f'Thresholded Image \t threshold: {self.threshold} \t circularity: {self.circ}')
            plt.tight_layout()
            figs.append(fig)

        with PdfPages(export) as pdf:
            for fig in figs:
                pdf.savefig(fig)
                plt.close(fig)

    def run(self):
        diff_img = self.getImage()
        round_contours = self.getContours(diff_img)
        x, y, radius = self.findLargestRound(round_contours)
        push_ = (self.path, radius*2/self.scale*10**3)
        dias = self.listAllAbove(round_contours)
        self.signals.push.emit(push_)
        self.signals.diasPush.emit(dias)
        self.signals.finished.emit()

class PDAWorkerSignals(QObject):
    finished = pyqtSignal() 
    push = pyqtSignal(list)

class PDAWorker(QRunnable):

    def __init__(self, path, upperCutOff, liqDens, row, mode='mat', matPath = None, scriptPath = None):
        super().__init__()
        self.path = path
        self.liqDens = liqDens
        self.upperCutOff = upperCutOff
        self.signals = PDAWorkerSignals()
        self.row = row
        self.mode = mode
        self.matPath = matPath
        self.scriptPath = scriptPath

    def run(self):
        pda = PDA.PDA(self.path, upperCutOff=self.upperCutOff, liqDens=self.liqDens, mode=self.mode, matPath=self.matPath, scriptPath = self.scriptPath)
        try:
            i, j, df_x = pda.run()
            self.signals.push.emit([i, j, self.row, df_x])
        except:
            self.signals.push.emit([None, None, None, None])
        self.signals.finished.emit()

class MatplotlibWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.figure, self.ax = plt.subplots()
        self.canvas = FigureCanvas(self.figure)
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(self.canvas)
        self.canvas.mpl_connect('button_press_event', self.get_coordinates)
        self.canvas.mpl_connect('scroll_event', self.on_scroll)
        self.l_cord = []
        self.r_cord = []
        self.lines = []
        self.leftPoint = None
        self.rightPoint = None

    def get_coordinates(self, event):
        if event.button == 1:
            x = event.xdata
            y = event.ydata

            for line in self.l_cord:
                line.remove()
                self.l_cord = []

            self.ax.scatter(x, y, color='red', marker='x')
            self.l_cord.append(self.ax.collections[-1])
            self.leftPoint = (x, y)

            self.canvas.draw()

        if event.button == 3:
            x = event.xdata
            y = event.ydata

            for line in self.r_cord:
                line.remove()
                self.r_cord = []

            self.ax.scatter(x, y, color='blue', marker='x')
            self.r_cord.append(self.ax.collections[-1])
            self.rightPoint = (x, y)

            self.canvas.draw()

    def on_scroll(self, event):
        if event.inaxes is not None:
            ax = event.inaxes

            xlim = ax.get_xlim()
            ylim = ax.get_ylim()

            x_center = event.xdata
            y_center = event.ydata

            zoom_factor = 0.9 if event.button == 'up' else 1.1

            new_xlim = (xlim[0] - x_center) * zoom_factor + x_center, (xlim[1] - x_center) * zoom_factor + x_center
            new_ylim = (ylim[0] - y_center) * zoom_factor + y_center, (ylim[1] - y_center) * zoom_factor + y_center

            ax.set_xlim(new_xlim)
            ax.set_ylim(new_ylim)

            self.canvas.draw()

    def update(self):
        self.canvas.draw()
    
    def clearPoints(self):
        try:
            for line in self.r_cord:
                line.remove()
                self.r_cord = []
        except: pass
        try: 
            for line in self.l_cord:
                line.remove()
                self.l_cord = []
        except: pass
        finally: 
            self.canvas.draw()
            self.leftPoint = None
            self.rightPoint = None

    def drawPoints(self):
        if self.leftPoint != None:
            self.l_cord = []
            self.ax.scatter(self.leftPoint[0], self.leftPoint[1], color='red', marker='x')
            self.l_cord.append(self.ax.collections[-1])
            self.canvas.draw()
            
        if self.rightPoint != None:
            self.r_cord = []
            self.ax.scatter(self.rightPoint[0], self.rightPoint[1], color='blue', marker='x')
            self.r_cord.append(self.ax.collections[-1])
            self.canvas.draw()

    def clear(self):
        try:
            self.canvas.axes.cla()
        except: pass
        self.canvas.draw()
    
    def reset(self):
        return
        self.figure, self.ax = plt.subplots()
        self.canvas = FigureCanvas(self.figure)

class AngleSignals(QObject):
    finished = pyqtSignal() 
    push = pyqtSignal(np.ndarray)

class AngleWorker(QRunnable):

    def __init__(self, path, ref):
        super().__init__()
        self.path = path
        self.ref = ref
        self.signals = AngleSignals()

    def run(self):
        read = mA.multiAngle(self.path, self.ref)
        # try:
        self.signals.push.emit(read.imageHandling())
        self.signals.finished.emit()
        # except:
        #     raise ValueError

# ---- End of processes ----        
# ---- Start of main programm ----

class UI(QMainWindow):
    def __init__(self):
        '''
        Initialization of the main programm.
        '''
        super().__init__()
        self.ui = main()
        self.ui.setupUi(self)
        if not os.path.exists(os.path.join(os.path.expanduser('~'), 'Atomizer Toolbox')): # Creation of a user settings folder. Try not to change the path, or make sure to change in packages as well if neccessary
            os.mkdir(os.path.join(os.path.expanduser('~'), 'Atomizer Toolbox'))
        self.path = os.path.join(os.path.expanduser('~'), 'Atomizer Toolbox')
        self.resultLabels = self.createLabelList()
        self.presetField()
        self.setWindowIcon(QtGui.QIcon('assets/ATT_LOGO.ico'))
        self.enableAutoCalc()
        self.initCalculator()
        # ---- Connecting button presses with corrosponding functions of the programm ---
        self.ui.cpToclip.clicked.connect(self.toClip)
        self.ui.cellToClip.clicked.connect(self.cellToClip)
        self.ui.actionEdit_and_Create_Export_Presets.triggered.connect(self.createExportPresets)
        self.ui.actionBulk_Generate.triggered.connect(self.bulkCalc)
        self.ui.actionSetup_Bulk_Export.triggered.connect(self.bulkExport)
        self.ui.actionLoad_Presets.triggered.connect(self.loadPreset)
        self.ui.actionSave_Presets.triggered.connect(self.savePreset)
        self.ui.actionReset_Values.triggered.connect(self.resetValues)
        self.ui.actionAbout.triggered.connect(self.about)
        self.ui.actionGo_to_default_path.triggered.connect(self.openPath)
        self.ui.loadInput.clicked.connect(self.loadInput)
        self.ui.selectFolderConverter.clicked.connect(self.convertFolder)
        self.ui.nextPic.clicked.connect(self.nextPic)
        self.ui.prevPic.clicked.connect(self.prevPic)
        self.ui.plus10.clicked.connect(self.next10Pic)
        self.ui.minus10.clicked.connect(self.prev10Pic)
        self.ui.lineDown.clicked.connect(self.moveLineDown)
        self.ui.lineUp.clicked.connect(self.moveLineUp)
        self.ui.line10Down.clicked.connect(self.moveLine10Down)
        self.ui.line10Up.clicked.connect(self.moveLine10Up)
        self.ui.picID.valueChanged.connect(self.load_image_into_graphics_view)
        self.ui.loadFolder.clicked.connect(self.loadFolder)
        self.cineItems = None
        self.ui.runConversion.clicked.connect(self.runConversion)
        self.ui.freqRun.clicked.connect(self.createFreqList)
        self.ui.clearRef.clicked.connect(self.showFFTArray)
        self.ui.clearRef.setDisabled(True)
        self.ui.loadRef.clicked.connect(self.loadRef)
        self.ui.dropFolder.clicked.connect(self.loadDropletFolder)
        self.ui.dropletRun.clicked.connect(self.dropletRun)
        self.ui.dropletRef.clicked.connect(self.loadDropletRef)
        self.ui.dropletGenerateReport.clicked.connect(self.generateReport)
        self.ui.PDA_load_Folder_1.clicked.connect(partial(self.loadPDAFolder, 0))
        self.ui.PDA_load_Folder_2.clicked.connect(partial(self.loadPDAFolder, 1))
        self.ui.PDA_load_Folder_3.clicked.connect(partial(self.loadPDAFolder, 2))
        self.ui.PDA_run.clicked.connect(self.runPDA)
        self.ui.PDA_D32.clicked.connect(self.createD32Graph)
        self.ui.PDA_vel.clicked.connect(self.createVelGraph)
        self.ui.PDA_Vel_mean.clicked.connect(self.createVelMeanGraph)
        self.ui.PDA_D32_mean.clicked.connect(self.createD32MeanGraph)
        self.ui.PDA_D32.setDisabled(True)
        self.ui.PDA_vel.setDisabled(True)
        self.ui.PDA_D32_mean.setDisabled(True)
        self.ui.PDA_Vel_mean.setDisabled(True)
        self.ui.runPat.clicked.connect(self.patternatorRun)
        self.ui.createExcel.clicked.connect(self.createExcel)
        # ---- End of button functions ----
        self.currentPDAFolder = None
        self.settings_dict = {}
        self.createFolders()
        self.lastMode = None
        self.lastFolder = None
        self.removePresetTag()
        self.tabOrder()
        self.loadGlobalSettings()
        self.loadStyles()
        self.checkMatlabInstalled()
        self.sprayAngleInit()
        self.patternatorInit()
        self.ui.angleRun.clicked.connect(self.angleRun)

# ---- Annular atomizer calculator from here ----

    def liqAndGasDF(self) -> pd.DataFrame: 
        '''
        Creation of dataframes with data for liquid and gas
        Returns pandas dataframe with the values
        '''
        liqAndGas = pd.DataFrame()
        liquidData = {
            'type': ['Liquid'],
            'Surface Tension [mN/m]': [self.liqSurface*1000],
            'Density [mg/m³]': [self.liqDens],
            'Viscosity [mPa s]': [self.liqVisc*1000],
            'Temperature [°C]': [self.ui.liquidTemp.value()]
        }
        newRow = pd.DataFrame(liquidData)
        liqAndGas = pd.concat([liqAndGas, newRow], ignore_index=True)

        gasData = {
        'type': ['Gas'],
        'Density [mg/m³]': [self.gasDens],
        'Temperature [°C]': [self.ui.gasTemp.value()]
        }
        newRow = pd.DataFrame(gasData)
        liqAndGas = pd.concat([liqAndGas, newRow], ignore_index=True)
        liqAndGas = liqAndGas.set_index('type')

        return liqAndGas

    def GeometryDF(self)->pd.DataFrame:
        '''
        Creation of dataframes with Geometry of atomizer
        Return pandas dataframe with data
        '''
        geometry = {
            'type': ['Geometry'],
            'Inner Tube Diameter [mm]': [self.innerTube*1000],
            'Inner Wall Thickness [mm]': [self.innerWall*1000],
            'Middle Sheet Thickness [mm]': [self.annularSheet*1000],
            'Middle Wall Thickness [mm]': [self.middleWall*1000],
            'Outer Sheet Thickness [mm]': [self.outerSheet*1000],
            'Inner Orifice [mm²]': [self.innerArea*1000**2],
            'Middle Orifice [mm²]': [self.liquidArea*1000**2],
            'Outer Orifice [mm²]': [self.outerArea*1000**2],
        }
        newRow = pd.DataFrame(data=geometry).set_index('type')
        return newRow

    def tabOrder(self) -> None:
        '''
        Initilized Tab order to use the tabulator key to move between input fields
        '''
        order = [self.ui.innerTube, self.ui.innerWall, self.ui.annularSheet, self.ui.middleWall, self.ui.outerSheet, self.ui.liquidTemp, self.ui.liquidVisc, self.ui.gasTemp, self.ui.innerStreamValue, self.ui.innerStreamUnit, self.ui.sheetStreamValue, self.ui.sheetStreamUnit, self.ui.outerStreamValue, self.ui.outerStreamUnit, self.ui.cpToclip]
        self.setTabOrder(order[0], order[1])
        self.setTabOrder(order[-1], order[0])
        for i in range(1, len(order)):
            self.setTabOrder(order[i-1], order[i])
        
        viscOrder = self.ui.input_my, self.ui.input_T, self.ui.viscoGlyPurity, self.ui.viscoGlyMass

        self.setTabOrder(viscOrder[0], viscOrder[1])
        self.setTabOrder(viscOrder[-1], viscOrder[0])
        for i in range(1, len(viscOrder)):
            self.setTabOrder(viscOrder[i-1], viscOrder[i])
    
    def setCalcButtons(self)-> None:
        '''
        Intilization of calculation functions for the Annular Atomzier Tab
        '''
        self.ui.liquidVisc.textChanged.connect(self.calcLiq)
        self.ui.liquidTemp.textChanged.connect(self.calcLiq)
        self.ui.gasTemp.textChanged.connect(self.calcGas)
        self.ui.liquidType.currentTextChanged.connect(self.calcLiq)

        self.calcLiq()
        self.calcGas()
        
    def enableAutoCalc(self) -> None:
        '''
        Initializes automatic calculation in the Annular Atomizer tab
        '''
        # self.ui.liquidTemp.valueChanged.connect(self.calcLiq)
        # self.ui.liquidVisc.valueChanged.connect(self.calcLiqDens)
        # self.ui.gasTemp.valueChanged.connect(self.calcGas)
        self.ui.liquidTemp.valueChanged.connect(self.readValues)
        self.ui.liquidVisc.valueChanged.connect(self.readValues)
        self.ui.gasTemp.valueChanged.connect(self.readValues)
        self.ui.innerTube.valueChanged.connect(self.readValues)
        self.ui.innerWall.valueChanged.connect(self.readValues)
        self.ui.annularSheet.valueChanged.connect(self.readValues)
        self.ui.middleWall.valueChanged.connect(self.readValues)
        self.ui.outerSheet.valueChanged.connect(self.readValues)
        self.ui.outerWall.valueChanged.connect(self.readValues)
        self.ui.innerStreamValue.valueChanged.connect(self.readValues)
        self.ui.sheetStreamValue.valueChanged.connect(self.readValues)
        self.ui.outerStreamValue.valueChanged.connect(self.readValues)
        self.ui.innerStreamUnit.currentTextChanged.connect(self.readValues)
        self.ui.sheetStreamUnit.currentTextChanged.connect(self.readValues)
        self.ui.outerStreamUnit.currentTextChanged.connect(self.readValues)

    def calcLiq(self)->bool:
        '''
        Calculation of liquid data like density, surface tension, ...
        Returns bool to check if calculation was successful, e.g. if data are out of bounds like viscosity of 9999 mPa s at 100 °C
        '''
       
        if self.ui.liquidType.currentText() == 'Water':
            water_ = Fluid(FluidsList.Water).with_state(Input.temperature(self.ui.liquidTemp.value()), Input.pressure(101325))
            self.ui.liquidVisc.setValue(water_.dynamic_viscosity*1000)
            self.liqDens = water_.density
            self.ui.LiquidDens.setValue(water_.density)
            self.liqSurface = water_ST = 235.8e-3*((water_.critical_temperature-self.ui.liquidTemp.value())/(water_.critical_temperature+273.15))**1.256*(1-0.625*((water_.critical_temperature-self.ui.liquidTemp.value())/(water_.critical_temperature+273.15)))
            self.ui.liquidVisc.setStyleSheet('')
            check =  True
        else: 
            try: 
                glycerinFraction = ca(self.ui.liquidTemp.value(), self.ui.liquidVisc.value()).solve()
                if glycerinFraction == ValueError: raise ValueError
            except ValueError: 
                glycerinFraction = 0
                # QMessageBox.warning(self, 'ERROR', 'This viscosity is not possible!')
            rhoGly = ca(self.ui.liquidTemp.value(), self.ui.liquidVisc.value()).rhoGlycerin()
            water_ = Fluid(FluidsList.Water).with_state(Input.temperature(self.ui.liquidTemp.value()), Input.pressure(101325))
            rhoMix = rhoGly*glycerinFraction + water_.density*(1-glycerinFraction)
            if glycerinFraction != 0: 
                self.ui.LiquidDens.setValue(rhoMix)
                self.ui.liquidVisc.setStyleSheet('')
                check =  True
            else: 
                self.ui.LiquidDens.setValue(1)
                self.ui.liquidVisc.setStyleSheet('background-color: red;')
                check = False
            self.liqDens = rhoMix
            surfaceGly = round(0.06*self.ui.liquidTemp.value()+64.6, 10)
            surfaceGly /= 1000
            
            water_ST = 235.8e-3*((water_.critical_temperature-self.ui.liquidTemp.value())/(water_.critical_temperature+273.15))**1.256*(1-0.625*((water_.critical_temperature-self.ui.liquidTemp.value())/(water_.critical_temperature+273.15)))
            surfaceMix = glycerinFraction*surfaceGly + (1-glycerinFraction)*water_ST
            self.liqSurface = surfaceMix
        return check
        
    def calcGas(self)->bool:
        '''
        Calculation of gas data like density, ...
        Returns bool to check if calculation was successful.
        '''
        try:
            if self.ui.gasType.currentText() == 'Air':
                air_ = Fluid(FluidsList.Air).with_state(Input.temperature(self.ui.gasTemp.value()), Input.pressure(101325))
                self.ui.gasVisc.setValue(air_.dynamic_viscosity*1000)
                self.ui.gasDens.setValue(air_.density)
                self.gasDens = air_.density
                self.gasSurface = air_.surface_tension
                return True
        except: 
            pass
            return False

    def createLabelList(self)->list:
        '''
        Initialize list of labels of the calculator to be used for filling them with values
        '''
        resultLabels = []
        resultLabels.append(self.ui.innerMFR)
        resultLabels.append(self.ui.sheetMFR)
        resultLabels.append(self.ui.outerMFR)

        resultLabels.append(self.ui.innerVel)
        resultLabels.append(self.ui.sheetVel)
        resultLabels.append(self.ui.outerVel)

        resultLabels.append(self.ui.innerMom_flux)
        resultLabels.append(self.ui.sheetMom_flux)
        resultLabels.append(self.ui.outerMom_flux)

        resultLabels.append(self.ui.innerMom)
        resultLabels.append(self.ui.sheetMom)
        resultLabels.append(self.ui.outerMom)

        resultLabels.append(self.ui.innerRe)
        resultLabels.append(self.ui.innerWe)
        resultLabels.append(self.ui.innerOh)
        self.ui.innerOh.setHidden(True)

        resultLabels.append(self.ui.sheetRe)
        resultLabels.append(self.ui.sheetWe)
        resultLabels.append(self.ui.sheetOh)
        self.ui.outerOh.setHidden(True)
    
        resultLabels.append(self.ui.outerRe)
        resultLabels.append(self.ui.outerWe)
        resultLabels.append(self.ui.outerOh)
        
        resultLabels.append(self.ui.totalGLR)
        resultLabels.append(self.ui.totalMom)
        resultLabels.append(self.ui.innerSheetGLR)
        resultLabels.append(self.ui.outerSheetGLR)
        resultLabels.append(self.ui.innerSheetMom)
        resultLabels.append(self.ui.outerSheetMom)
        resultLabels.append(self.ui.weberGLR)

        for label in resultLabels:
            label.setText('-')
        return resultLabels

    # readValues(self) is important and handles the most tasks
    def readValues(self)-> None:
        '''
        Read current input and runs calculation
        '''
        if not self.calcLiq() or not self.calcGas(): return 0
        # Atomizer Geometry
        self.Lc = []
        self.innerTube = self.ui.innerTube.value()/1000
        self.Lc.append(self.innerTube)
        self.innerWall = self.ui.innerWall.value()/1000
        self.annularSheet = self.ui.annularSheet.value()/1000
        self.Lc.append(2*self.annularSheet)
        self.middleWall = self.ui.middleWall.value()/1000
        self.outerSheet = self.ui.outerSheet.value()/1000
        self.Lc.append(2*self.outerSheet)
        self.outerWall = self.ui.outerWall.value()/1000
        self.orifice()

        check = [self.innerTube, self.annularSheet, self.outerSheet]
        for item in check: 
            if item == 0: return 0

        # Fluid Properties
        ## Liquid
        self.liqType = self.ui.liquidType.currentText()
        self.liqTemp = self.ui.liquidTemp.value()+273.15
        self.liqVisc = self.ui.liquidVisc.value()/1000
        self.liqDens = self.ui.LiquidDens.value()

        ## Gas
        self.gasType = self.ui.gasType.currentText()
        self.gasTemp = self.ui.gasTemp.value()+275.15
        self.gasVisc = self.ui.gasVisc.value()/1000
        self.gasDens = self.ui.gasDens.value()

        # Stream Properties
        # values are converted to kg/s
        self.streams = []
        value = self.ui.innerStreamValue.value()
        unit = self.ui.innerStreamUnit.currentText()
        if self.ui.innerGasTrue.isChecked():
            type = 'gas'
            self.ui.innerOh.setHidden(True)
        else: 
            type = 'liquid'
            self.ui.innerOh.setHidden(False)
        self.streams.append([value, unit, type])

        value = self.ui.sheetStreamValue.value()
        unit = self.ui.sheetStreamUnit.currentText()
        if self.ui.sheetGasTrue.isChecked():
            type = 'gas'
            self.ui.sheetOh.setHidden(True)
        else: 
            type = 'liquid'
            self.ui.sheetOh.setHidden(False)
        self.streams.append([value, unit, type])

        value = self.ui.outerStreamValue.value()
        unit = self.ui.outerStreamUnit.currentText()
        if self.ui.outerGasTrue.isChecked():
            type = 'gas'
            self.ui.outerOh.setHidden(True)
        else: 
            type = 'liquid'
            self.ui.outerOh.setHidden(False)
        self.streams.append([value, unit, type])

        self.streamValues = []
        def velCalc(mfr, orifice, type):
            mfr_h = mfr*3600
            if type == 'gas':
                dens = self.gasDens 
            else: dens = self.liqDens
            vel = mfr/dens/orifice
            mom = dens*vel**2*orifice
            mom_flux = dens*vel**2
            return [mfr, mfr_h, vel, mom_flux, mom, type]

        i = 0
        for item in self.streams:
            area = self.orificeDict[i]
            i += 1
            value, unit, type = item
            if unit == 'kg/h':
                value /= 3600
            elif unit == 'g/s':
                value /= 1000
            elif unit == 'm/s':
                value *= area
                if type == 'gas':
                    value *= self.gasDens
                else:
                    value *= self.liqDens 

            self.streamValues.append(velCalc(value, area, type))

        self.fillFirstResults()

    def orifice(self)->None:
        '''
        Calculation of the orifices of the atomizer
        '''
        self.innerArea = math.pi/4*self.innerTube**2
        self.innerAreaWithWall = math.pi*(self.innerTube/2+self.innerWall)**2

        liquidSheetRadius = self.innerTube/2 + self.annularSheet + self.innerWall
        self.liquidArea = math.pi*liquidSheetRadius**2 - self.innerAreaWithWall
        self.liquidAreaWithWall = math.pi*(liquidSheetRadius+self.middleWall)**2 - self.innerAreaWithWall

        outerSheetRadius = liquidSheetRadius + self.middleWall + self.outerSheet
        self.outerArea = math.pi*outerSheetRadius**2 - self.liquidAreaWithWall - self.innerAreaWithWall
        self.orificeDict = {
            0: self.innerArea,
            1: self.liquidArea,
            2: self.outerArea
        }
        # print(self.orificeDict)

    def fillFirstResults(self)->None:
        '''
        Inputs the calculated values to the labels
        '''
        streamValuesString = [[0]*5 for i in range(3)]
        for j in range(len(self.streamValues[0])-1):
            for i in range(3):
                if self.streamValues[i][j] < 0.01:
                    streamValuesString[i][j] = "%.3e" % self.streamValues[i][j]
                elif self.streamValues[i][j] > 1000:
                    streamValuesString[i][j] = "%.3e" % self.streamValues[i][j]
                else: 
                    streamValuesString[i][j] = "%.3f" % self.streamValues[i][j]
            

        self.resultLabels[0].setText(streamValuesString[0][1])
        self.resultLabels[1].setText(streamValuesString[1][1])
        self.resultLabels[2].setText(streamValuesString[2][1])

        self.resultLabels[3].setText(streamValuesString[0][2])
        self.resultLabels[4].setText(streamValuesString[1][2])
        self.resultLabels[5].setText(streamValuesString[2][2])
        
        self.resultLabels[6].setText(streamValuesString[0][3])
        self.resultLabels[7].setText(streamValuesString[1][3])
        self.resultLabels[8].setText(streamValuesString[2][3])
        
        self.resultLabels[9].setText(streamValuesString[0][4])
        self.resultLabels[10].setText(streamValuesString[1][4])
        self.resultLabels[11].setText(streamValuesString[2][4])

        def StreamDF()->pd.DataFrame:
            '''
            Sets up multiple dataframes with current values, which are the used with the export of calculated data
            '''
            dict = {
                'type': ['inner Stream', 'middle Stream', 'outer Stream'],
                'Fluid Type': [self.streamValues[0][-1].capitalize(), self.streamValues[1][-1].capitalize(), self.streamValues[2][-1].capitalize()],
                'Mass Flow Rate [kg/h]': [self.streamValues[0][1], self.streamValues[1][1], self.streamValues[2][1]],
                'Flow Velocity [m/s]': [self.streamValues[0][2], self.streamValues[1][2], self.streamValues[2][2]],
                'Momentum Flux [kg/(m s²)]': [self.streamValues[0][3], self.streamValues[1][3], self.streamValues[2][3]],
                'Momentum [kg m/s²]': [self.streamValues[0][4], self.streamValues[1][4], self.streamValues[2][4]]
            }
            
            return pd.DataFrame(dict).set_index('type')

        self.StreamDf = StreamDF()
        self.calcDimless()

    def calcDimless(self)-> None:
        '''
        Calculation of diml. numbers (Re, Oh, We, ...)
        '''
        rhos = {
            'gas': self.gasDens,
            'liquid': self.liqDens
        }
        visc = {
            'gas': self.gasVisc,
            'liquid': self.liqVisc
        }
        sigmas = {
            'gas': self.liqSurface,
            'liquid': self.liqSurface
        }
        
        self.innerDimless = []
        self.sheetDimless = []
        self.outerDimless = []

        self.innerDimless.append(dL.Re(self.streamValues[0][2], rhos[self.streamValues[0][5]], self.Lc[0], visc[self.streamValues[0][5]]))
        self.sheetDimless.append(dL.Re(self.streamValues[1][2], rhos[self.streamValues[1][5]], self.Lc[1], visc[self.streamValues[1][5]]))
        self.outerDimless.append(dL.Re(self.streamValues[2][2], rhos[self.streamValues[2][5]], self.Lc[2], visc[self.streamValues[2][5]]))
        
        self.innerDimless.append(dL.We(self.streamValues[0][2], rhos[self.streamValues[0][5]], self.Lc[0], sigmas[self.streamValues[0][5]]))
        self.sheetDimless.append(dL.We(self.streamValues[1][2], rhos[self.streamValues[1][5]], self.Lc[1], sigmas[self.streamValues[1][5]]))
        self.outerDimless.append(dL.We(self.streamValues[2][2], rhos[self.streamValues[2][5]], self.Lc[2], sigmas[self.streamValues[2][5]]))
        
        self.innerDimless.append(dL.Oh(visc[self.streamValues[0][5]], rhos[self.streamValues[0][5]], self.Lc[0], sigmas[self.streamValues[0][5]]))
        self.sheetDimless.append(dL.Oh(visc[self.streamValues[1][5]], rhos[self.streamValues[1][5]], self.Lc[1], sigmas[self.streamValues[1][5]]))
        self.outerDimless.append(dL.Oh(visc[self.streamValues[2][5]], rhos[self.streamValues[2][5]], self.Lc[2], sigmas[self.streamValues[2][5]]))

        try:
            self.We_GLR = dL.We_GLR(rhos[self.streamValues[1][5]], sigmas[self.streamValues[1][5]], self.Lc[1], self.streamValues[0][1]/self.streamValues[1][1], 
                                    self.streamValues[2][1]/self.streamValues[1][1], self.streamValues[0][2], self.streamValues[2][2])
        except: self.We_GLR = 0

        def ReWeOhDF()->pd.DataFrame:
            '''
            Creates a dict of dataframes with diml. numbers for each stream
            '''
            dict = {
                'type': ['inner Stream', 'middle Stream', 'outer Stream'], 
                'Reynolds': [self.innerDimless[0], self.sheetDimless[0], self.outerDimless[0]],
                'Weber': [self.innerDimless[1], self.sheetDimless[1], self.outerDimless[1]],
                'Ohnesorge': [self.innerDimless[2], self.sheetDimless[2], self.outerDimless[2]],
                'Weber GLR': [None, self.We_GLR, None]
            }

            return pd.DataFrame(dict).set_index('type')

        self.ReOhWeDf = ReWeOhDF()

        # --- Creation of strings for labels ---
        # Labels have to be filled with strings and not numbers directly

        strings = []
        for item in self.innerDimless:
            if item < 0.01 or item > 1000:
                strings.append("%.2e" % item)
            else: 
                strings.append("%.2f" % item)
        
        for item in self.sheetDimless:
            if item < 0.01 or item > 1000:
                strings.append("%.2e" % item)
            else: 
                strings.append("%.2f" % item)
        
        for item in self.outerDimless:
            if item < 0.01 or item > 1000:
                strings.append("%.2e" % item)
            else: 
                strings.append("%.2f" % item)

        for i in range(len(strings)):
            self.resultLabels[12+i].setText(strings[i])
        
        if self.We_GLR > 1000:
            self.resultLabels[27].setText("%.3e" % self.We_GLR)
        else: 
            self.resultLabels[27].setText("%.2f" % self.We_GLR)

        # GLR and momentum flux calculation 

        gasMF = 0
        liqMF = 0
        gasMom_flux = 0
        liqMom_flux = 0
        
        for i in range(3):
            if self.streamValues[i][-1] == 'gas':
                gasMF += self.streamValues[i][0]
                gasMom_flux += self.streamValues[i][3]
            elif self.streamValues[i][-1] == 'liquid':
                liqMF += self.streamValues[i][0]
                liqMom_flux += self.streamValues[i][3]

        #
        self.RatiosDf = pd.DataFrame(columns=['GLR', 'GLI', 'GLO', 'Momentum Flux Ratio', 'Inner Momentum Flux Ratio', 'Outer Momentum Flux Ratio'])

        if liqMF != 0 and liqMom_flux != 0:
            self.GLR_total = gasMF/liqMF
            self.mom_flux_total = gasMom_flux/liqMom_flux
            if self.streamValues[0][-1] == 'gas' and self.streamValues[1][-1] == 'liquid' and self.streamValues[2][-1] == 'gas':
                self.GLI = self.streamValues[0][0]/self.streamValues[1][0]
                self.GLO = self.streamValues[2][0]/self.streamValues[1][0]
                self.mom_flux_i = self.streamValues[0][3]/self.streamValues[1][3]
                self.mom_flux_o = self.streamValues[2][3]/self.streamValues[1][3]
                self.mom_i = self.streamValues[0][4]/self.streamValues[1][4]
                self.mom_o = self.streamValues[2][4]/self.streamValues[1][4]
                self.mom_tot = (self.streamValues[0][4]+self.streamValues[2][4])/self.streamValues[1][4]

                dict = {
                    'GLR': [self.GLR_total],
                    'GLI': [self.GLI],
                    'GLO': [self.GLO],
                    'Total Mometum Ratio J': [self.mom_tot],
                    'Inner Mometum Ratio J_i': [self.mom_i],
                    'Outer Mometum Ratio J_o': [self.mom_o],
                    'Momentum Flux Ratio': [self.mom_flux_total],
                    'Inner Momentum Flux Ratio': [self.mom_flux_i],
                    'Outer Momentum Flux Ratio': [self.mom_flux_o],
                    'Total Gas Momentum [kg m/s²]': [self.streamValues[0][4]+self.streamValues[2][4]]
                }
                self.RatiosDf = pd.DataFrame(dict)
            else: 
                self.GLI = 0
                self.GLO = 0
                self.mom_flux_i = 0
                self.mom_flux_o = 0
                  
            self.resultLabels[21].setText("%.2f" % self.GLR_total)
            self.resultLabels[22].setText("%.2f" % self.mom_tot)
            self.resultLabels[23].setText("%.2f" % self.GLI)
            self.resultLabels[24].setText("%.2f" % self.GLO)
            self.resultLabels[25].setText("%.2f" % self.mom_i)
            self.resultLabels[26].setText("%.2f" % self.mom_o)

        else:
            self.resultLabels[21].setText('Error')
            self.resultLabels[22].setText('Error')
            self.resultLabels[23].setText('Error')
            self.resultLabels[24].setText('Error')
            self.resultLabels[25].setText('Error')
            self.resultLabels[26].setText('Error')
            

            dict = {
                    'GLR': ['--'],
                    'GLI': ['--'],
                    'GLO': ['--'],
                    'Total Mometum Ratio J': ['--'],
                    'Inner Mometum Ratio J_i': ['--'],
                    'Outer Mometum Ratio J_o': ['--'],
                    'Momentum Flux Ratio': ['--'],
                    'Inner Momentum Flux Ratio': ['--'],
                    'Outer Momentum Flux Ratio': ['--'],
                    'Total Gas Momentum [kg m/s²]': ['--']
                }
            self.RatiosDf = pd.DataFrame(dict)

        self.getAllDfs()

    # --- Handling of export of data to excel or clipboard ---

    def createFolders(self)-> None:
        '''
        Creates hidden export folders in the user folder to handle the data of the export
        '''
        default = {
                'lastFile': 'empty__',
                'lastExport': 'empty__', 
                'exportDecimal': 'point', 
                'exportHeader': False
                }
        
        if not os.path.exists(os.path.join(self.path, 'global')): 
            os.mkdir(os.path.join(self.path, 'global'))
            if not os.path.exists(os.path.join(self.path, 'global', 'share')):
                os.mkdir(os.path.join(self.path, 'global', 'share'))

            FILE_ATTRIBUTE_HIDDEN = 0x02
            ctypes.windll.kernel32.SetFileAttributesW(fr"{os.path.join(self.path, 'global')}", FILE_ATTRIBUTE_HIDDEN)
            self.settings = settings(os.path.join(self.path, 'global', 'global_settings.json'))
            self.exportDB = exportDB(os.path.join(self.path, 'global', 'export.db'))
            self.settings.setup(default)
            self.resetValues()
        else:
            self.settings = settings(os.path.join(self.path, 'global', 'global_settings.json'))
            self.exportDB = exportDB(os.path.join(self.path, 'global', 'export.db'))

    def loadGlobalSettings(self)->None:
        '''
        Loads user settings
        '''
        path = os.path.join(self.path, 'global', 'global_settings.json')
        if not os.path.exists(path):
            try: 
                default = {
                    'lastFile': 'empty__',
                    'lastExport': 'empty__', 
                    'exportDecimal': 'point', 
                    'exportHeader': False, 
                    'PDA_auto_folder': True,
                    'PDA_full_export': True,
                    'PDA_diagrams': 'False',
                    'SpA_norm': 'True',
                    'SpA_maxW': 'False',
                    'patChan': 114, 
                    'patA': 5.5, 
                    'patB': 5.5, 
                    'patH': 0.35, 
                    'patOpTotal': 50, 
                    'patOp': 3, 
                }
                if not os.path.exists(os.path.join(self.path, 'global')): os.mkdir(os.path.join(self.path, 'global'))
                if not os.path.exists(os.path.join(self.path, 'global', 'share')):
                    os.mkdir(os.path.join(self.path, 'global', 'share'))

                FILE_ATTRIBUTE_HIDDEN = 0x02
                ctypes.windll.kernel32.SetFileAttributesW(fr"{os.path.join(self.path, 'global')}", FILE_ATTRIBUTE_HIDDEN)
                self.settings.setup(default)
                self.resetValues()
            except: pass
        else:
            # Load export decimal comma or point for export
            try:
                mode = self.settings.get('exportDecimal')
                if mode == 'point': self.ui.radioPoint.setChecked(True)
                else: self.ui.radioComma.setChecked(True)
            except:
                self.settings.set('exportDecimal', 'point')
                self.ui.radioPoint.setChecked(True)
            
            # Load export header setting
            try:
                if self.settings.get('exportHeader'):
                    self.ui.headerCheck.setChecked(True)
                else: self.ui.headerCheck.setChecked(False)
            except:
                self.settings.set('exportHeader', False)
                self.ui.headerCheck.setChecked(False)

            # Load preset
            try: self.loadPreset(self.settings.get('lastFile'))
            except: pass

            # Load PDA auto folder setting
            try: 
                if self.settings.get('PDA_auto_folder'): self.ui.actionAutomatic_Folder_Detection.setChecked(True)
                else: self.ui.actionAutomatic_Folder_Detection.setChecked(False)
            except: pass
            
            try: 
                if self.settings.get('PDA_full_export'): self.ui.actionGenerate_Full_Export.setChecked(True)
                else: self.ui.self.ui.actionGenerate_Full_Export.setChecked(False)
            except: pass
            
            try: 
                if self.settings.get('PDA_diagrams'): self.ui.actionGenerate_and_save_Diagrams.setChecked(True)
                else: self.ui.self.ui.actionGenerate_and_save_Diagrams.setChecked(False)
            except: pass
            
            try: 
                if self.settings.get('SpA_norm'): self.ui.actionNormalized_Propabilty_Map.setChecked(True)
                else: self.ui.actionNormalized_Propabilty_Map.setChecked(False)
            except: pass
            
            try: 
                if self.settings.get('SpA_maxW'): self.ui.actionMax_Width_Mode.setChecked(True)
                else: self.ui.actionMax_Width_Mode.setChecked(False)
            except: pass
        
    def loadPreset(self, path = None):
        '''
        Load atomizer presets if existing
        '''
        if not path: filename, null = QFileDialog.getOpenFileName(self, directory=self.path, filter='*.json', options=QFileDialog.Option.ReadOnly)
        elif path == 'empty__': return 0
        else: filename = path

        try: 
            with open(rf'{filename}', 'r') as read:
                import_dict = json.load(read)
        except: 
            print('File not found')
            return 0
        
        for i in range(len(self.presetFields)):
            self.presetFields[i].setValue(import_dict[f"{i}"])
        
        for i in range(len(self.presetDropdowns)):
            self.presetDropdowns[i].setCurrentIndex(import_dict[f"{len(self.presetFields)+i}"])
        
        for i in range(len(self.presetRadio)):
            self.presetRadio[i].setChecked(import_dict[f"{len(self.presetFields)+ len(self.presetDropdowns) +i}"])

        list = filename.split('/')
        self.ui.label_19.setText(f'Atomizer Properties (Preset: {list[-1][:-5]})')

        self.settings.set('lastFile', filename)

    def savePreset(self):
        '''
        Save presets
        '''
        filename, null = QFileDialog.getSaveFileName(self, directory=self.path, filter='*.json')

        export = {}
        i = 0
        for item in self.presetFields:
            export[i] = item.value()
            i += 1
            
        for item in self.presetDropdowns:
            export[i] = item.currentIndex()
            i += 1

        for item in self.presetRadio:
            export[i] = item.isChecked()
            i += 1

        self.export = export
        if not filename:
            return 0

        with open(rf'{filename}', 'w+') as json_file:
            json.dump(export, json_file)

        self.settings.set('lastFile', filename)

    def presetField(self):
        '''
        Setup of empty presets
        '''
        fields = [self.ui.innerTube, self.ui.innerWall, self.ui.annularSheet, self.ui.middleWall, self.ui.outerSheet, self.ui.outerWall, self.ui.liquidTemp, self.ui.liquidVisc, self.ui.gasTemp, self.ui.innerStreamValue, self.ui.sheetStreamValue, self.ui.outerStreamValue]
        defaults = [3, 0.1, 1, 0.1, 1, 0, 20, 1, 20, 0, 0, 0]
        self.presetFields = fields
        self.presetFieldsDeafault = defaults

        dropDowns = [self.ui.innerTubeBox, self.ui.innerWallBox, self.ui.annularSheetBox, self.ui.middleWallBox, self.ui.outerSheetBox, self.ui.outerWallBox, self.ui.liquidType, self.ui.gasType, self.ui.innerStreamUnit, self.ui.sheetStreamUnit, self.ui.outerStreamUnit]
        defaults = [0 for i in range(len(dropDowns))]
        self.presetDropdowns = dropDowns
        self.presetDropdownsDefault = defaults

      
        radio = [self.ui.innerGasTrue, self.ui.innerLiqTrue, self.ui.sheetGasTrue, self.ui.sheetLiqTrue, self.ui.outerGasTrue, self.ui.outerLiqTrue]
        defaults = [True, False, False, True, True, False]
        self.presetRadio = radio    
        self.presetRadioDefault = defaults

    def resetValues(self):
        '''
        Resets label fields to the default value
        '''
        for i in range(len(self.presetFields)):
            self.presetFields[i].setValue(self.presetFieldsDeafault[i])
        
        for i in range(len(self.presetDropdowns)):
            self.presetDropdowns[i].setCurrentIndex(self.presetDropdownsDefault[i])
        
        for i in range(len(self.presetRadio)):
            self.presetRadio[i].setChecked(self.presetRadioDefault[i])

        self.calcLiq()
        self.calcGas()

    def removePresetTag(self):
        fields = [self.ui.innerTube, self.ui.innerWall, self.ui.annularSheet, self.ui.middleWall, self.ui.outerSheet, self.ui.outerWall]
        for item in fields:
            item.valueChanged.connect(self.removeTag)
   
    def removeTag(self):
            self.ui.label_19.setText('Atomizer Properties')

    def getAllDfs(self):
        dfs = {
            'Liquid and Gas Properties': self.liqAndGasDF(),
            'Atomizer Geometry': self.GeometryDF(),
            'Stream Properties': self.StreamDf,
            'Dimensionless Numbers': self.ReOhWeDf,
            'Common Ratios': self.RatiosDf
        }

        if not os.path.exists(os.path.join(self.path, 'global', 'share')):
            os.mkdir(os.path.join(self.path, 'global', 'share'))
        if not os.path.exists(os.path.join(self.path, 'global', 'presets')):
            os.mkdir(os.path.join(self.path, 'global', 'presets'))

        for k,v in dfs.items():
            self.exportDB.writeData(v, k)
            with open(os.path.join(self.path, 'global', 'share', f'{k}_share.json'), 'w+') as file:
                v.to_json(file, default_handler=float)
            # print('\n')
          
    def loadStyles(self):
        self.ui.exportStyleBox.clear()
        
        self.ui.exportStyleBox.addItem('Select Export Style')
        if not os.path.exists(os.path.join(self.path, 'global', 'presets')):
            os.mkdir(os.path.join(self.path, 'global', 'presets'))
        items = os.listdir(os.path.join(self.path, 'global', 'presets'))
        self.ui.exportStyleBox.addItems(items)
        if self.settings.get('lastExport') != 'empty__':
            if self.settings.get('lastExport') in items:
                self.ui.exportStyleBox.setCurrentText(self.settings.get('lastExport'))

    def generateExport(self):
        style = self.ui.exportStyleBox.currentText()
        if style == 'Select Export Style':
            return None
        else:
            self.settings.set('lastExport', style)

            path = os.path.join(self.path, 'global', 'presets', style, 'database.db')
            ex = exportDB(path)
            dicts = ex.DBtoDicts()

            if not 'self.innerTube' in locals() or not 'self.innerTube' in globals():
                self.readValues()

            dfs = {
                'Liquid and Gas Properties': self.liqAndGasDF(),
                'Atomizer Geometry': self.GeometryDF(),
                'Stream Properties': self.StreamDf,
                'Dimensionless Numbers': self.ReOhWeDf,
                'Common Ratios': self.RatiosDf
            }

            df = pd.DataFrame(columns=dicts[1])
            for key, value in dicts[0].items():
                innerList = []
                for k, v in value.items():
                    if type(v) == list:
                        innerList.append(dfs[v[0]][v[1]][v[2]])
                    elif type(v) == None:
                        innerList.append(None)
                    else: 
                        innerList.append(v)
                df.loc[len(df)] = innerList

            vheader = [v for k,v in dicts[2].items()]
            df[''] = pd.Series(vheader).values
            df = df.set_index('')
            df = df.fillna('')
            return df

    def generateCells(self):
        style = self.ui.exportStyleBox.currentText()
        if style == 'Select Export Style':
            return None
        self.settings.set('lastExport', style)
        path = os.path.join(self.path, 'global', 'presets', style, 'database.db')
        ex = exportDB(path)
        dicts = ex.DBtoDicts()

        if not 'self.innerTube' in locals() or not 'self.innerTube' in globals():
            self.readValues()

        dfs = {
            'Liquid and Gas Properties': self.liqAndGasDF(),
            'Atomizer Geometry': self.GeometryDF(),
            'Stream Properties': self.StreamDf,
            'Dimensionless Numbers': self.ReOhWeDf,
            'Common Ratios': self.RatiosDf
        }
       
        df = pd.DataFrame(columns=dicts[1])
        for key, value in dicts[0].items():
            innerList = []
            for k, v in value.items():
                if type(v) == list:
                    innerList.append(f"{v[1]} : {v[2]}")
                elif type(v) == None:
                    innerList.append(None)
                else: 
                    innerList.append(v)
            df.loc[len(df)] = innerList

        vheader = [v for k,v in dicts[2].items()]
        df[''] = pd.Series(vheader).values
        df = df.set_index('')
        df = df.fillna('')
        return df

    def toClip(self):
        df = self.generateExport()
        if type(df) == type(pd.DataFrame()):
            df = self.replace(df)
            if self.ui.headerCheck.isChecked() == True:
                self.settings.set('exportHeader', True)
                df.to_clipboard(sep='\t')
            else: 
                df.to_clipboard(header=False, index=False, sep='\t')
                self.settings.set('exportHeader', False)
            self.changeColor(self.ui.cpToclip, 'green', 2000)
        else:
            self.changeColor(self.ui.exportStyleBox, 'red', 1000)
            return None

    def cellToClip(self):
        df = self.generateCells()
        if type(df) == type(pd.DataFrame()):
            if self.ui.headerCheck.isChecked() == True:
                df.to_clipboard(decimal=',', sep='\t')
                self.settings.set('exportHeader', True)
            else: 
                df.to_clipboard(header=False, index=False, decimal=',', sep='\t')
                self.settings.set('exportHeader', False)
            self.changeColor(self.ui.cellToClip, 'green', 1000)
        else:
            self.changeColor(self.ui.exportStyleBox, 'red', 1000)
            return None
        
    def replace(self, df):
        if self.ui.radioComma.isChecked() == True:
            self.settings.set('exportDecimal', 'comma')
            df = df.applymap(lambda x: str(x).replace('.', ','))
        else: 
            self.settings.set('exportDecimal', 'point')
            df = df.applymap(lambda x: str(x))
        # print(df)
        return df
    
    @staticmethod
    def DfToClip(df:pd.DataFrame, deciaml:str = ',', sep:str = '\t', index:bool = True, header:bool = True):
        cols = df.columns
        rows = df.index
        df.fillna('')

        df = df.applymap(lambda x: str(x).replace('.', ',') if x.isnumeric else x)

        if header: lines = [cols.values]
        else: lines = []
        i = 0
        for item in df.iterrows():
            line = []
            if index: line.append(rows.values[i])
            i += 1
            for i in range(item.__len__()):
                line.append(item[1].iloc[i])
            lines.append(line)

        out = ''
        for item in lines:
            for i in item:
                out += f'{i}{sep}'
            out += '\n'
        
        out = out[:-2]
        pc.copy(out)
        return out

    def changeColor(self, button, color, duration):
        self.color_timer = QTimer()
        self.color_timer.timeout.connect(partial(self.resetColor, button))
        button.setStyleSheet(f'background-color: {color};')
        self.color_timer.start(duration)  # 2000 milliseconds (2 seconds)

    def resetColor(self, button):
        button.setStyleSheet('')
        self.color_timer.stop()

    def createExportPresets(self):
        ui = ex.UI(self)
        ui.exec()
        self.loadStyles()

    def about(self):
        QMessageBox.information(self, 'About', 'Created by David Märker <br> <a href="https://github.com/realize-1337">Github Profile</a>')

    def openPath(self):
        subprocess.Popen(rf'explorer /select,"{self.path}"')

    def bulkCalc(self):
        if not os.path.exists(os.path.join(self.path, 'global', 'export', 'bulk.json')): return
        with open(os.path.join(self.path, 'global', 'export', 'bulk.json'), 'r') as file:
            data = json.load(file)
        inner = data['inner']
        liq = data['middle']
        outer = data['outer']
        try: 
            file = data['export'].replace('/', '\\')
            if '/' in file:
                file.replace('/', '\\')
        except: 
            QMessageBox.information(self, 'Error', f'Make sure to set an output')
            return
        self.ui.innerStreamUnit.setCurrentText(data['innerUnit'])
        self.ui.sheetStreamUnit.setCurrentText(data['middleUnit'])
        self.ui.outerStreamUnit.setCurrentText(data['outerUnit'])
        df = pd.DataFrame()
        for l in liq:
            for i in inner:
                for o in outer:
                    self.ui.innerStreamValue.setValue(i)
                    self.ui.sheetStreamValue.setValue(l)
                    self.ui.outerStreamValue.setValue(o)
                    self.readValues()
                    df0 = self.generateExport()
                    if type(df) == type(pd.DataFrame()):
                        # df0:pd.DataFrame = self.replace(df0)
                        df = pd.concat([df, df0])
                    else:
                        self.changeColor(self.ui.exportStyleBox, 'red', 1000)
                        return
        try: 
            with pd.ExcelWriter(file, mode='w') as writer:
                df.to_excel(writer)
                df = self.replace(df)
                df.to_clipboard(header=False, index=False, decimal=',', sep='\t')
        except PermissionError: 
            QMessageBox.information(self, 'Error', f'Make sure to close {file}')
        except: 
            self.changeColor(self.ui.exportStyleBox, 'red', 1000)
            return None
        self.changeColor(self.ui.exportStyleBox, 'green', 1000)

    def bulkExport(self):
        self.bulk = bulkex.UI(self)
        self.bulk.exec()

    # --- End of Annular Atomizer Tab ---
    # --- Start of ViscoCalc Tab ---

    def initCalculator(self):
        list = [
            self.ui.input_T,
            self.ui.input_my,
            self.ui.viscoGlyMass, 
            self.ui.viscoGlyPurity
        ]

        for l in list:
            l.valueChanged.connect(self.calculator)
        self.calculator()

    def calculator(self):
        self.ui.outputLabel.setText('Berechnung läuft')
        my_target = self.ui.input_my.value()
        temp = self.ui.input_T.value()
        purity = self.ui.viscoGlyPurity.value()
        glyMass = self.ui.viscoGlyMass.value()

        calc = ca(temp, my_target)
        try: 
            result_num = calc.solve()
            result = "%.2f" % (result_num*100)
            output = f'<u>{result} wt-%</u> of 100 % glycerin are requiered to get a viscosity of {"%.0f" % my_target} mPa s at {temp} °C.'
            if glyMass != 0:
                water = (glyMass*purity/100-glyMass*result_num)/result_num
                output += f'<br> For a glycerin mass of {glyMass} g of purity {purity} %, <u>{"%.2f" % water} g</u> is required.'
        except: output = '<br><br>This combination of viscosity and temperature is not possible!'
        self.ui.outputLabel.setText(output)

    # FREQUENCY ANALYSIS

    def nextPic(self):
        self.ui.picID.setValue(self.ui.picID.value()+1)
        self.checkButtonState()
        self.load_image_into_graphics_view()

    def prevPic(self):
        self.ui.picID.setValue(self.ui.picID.value()-1)
        self.checkButtonState()
        self.load_image_into_graphics_view()

    def next10Pic(self):
        self.ui.picID.setValue(self.ui.picID.value()+10)
        self.checkButtonState()
        self.load_image_into_graphics_view()

    def prev10Pic(self):
        self.ui.picID.setValue(self.ui.picID.value()-10)
        self.checkButtonState()
        self.load_image_into_graphics_view()

    def checkButtonState(self):
        if not self.lastMode: return
        elif self.lastMode == 'cine':
            count = 0
            for root, dir, files in os.walk(os.path.join(self.path, 'global', 'currentCine')):
                count += len(files)
            self.ui.picID.setMaximum(count-1)
        else: return

    def load_image_into_graphics_view(self):
        if not self.lastMode: return
        elif self.lastMode == 'cine': self.load_cine_into_graphics_view()
        else: self.load_folder_into_graphics_view(self.currentPathFreq)
        
    def load_cine_into_graphics_view(self):
        scene = QGraphicsScene()
        id = self.ui.picID.value()
        pixmap = QPixmap(os.path.join(self.path, 'global', 'currentCine', f'frame_{"%04d" % id}.jpeg'))

        if not pixmap.isNull():
            item = scene.addPixmap(pixmap)
            self.ui.graphicsView.setScene(scene)
            # self.ui.graphicsView.fitInView(item, aspectRatioMode=1)
            self.ui.graphicsView.fitInView(item)
        else:
            print("Failed to load image.")

        # pen = QPen(QColor(0, 135, 108)) # KIT COLOR
        pen = QPen(QColor(0, 0, 255)) # BLACK
        self.line_y = 100
        self.line = scene.addLine(0, self.line_y, pixmap.width(), 100, pen)
        self.line.setZValue(1) 

        self.line_x1 = 0
        self.line_x2 = pixmap.width()
        self.linex1 = scene.addLine(self.line_x1, 0, self.line_x1, pixmap.height(), pen)
        self.linex2 = scene.addLine(self.line_x2, 0, self.line_x2, pixmap.height(), pen)
        self.linex1.setZValue(1)
        self.linex2.setZValue(1)

    def load_folder_into_graphics_view(self, path, name='frame_', type='.png'):
        scene = QGraphicsScene()
        id = self.ui.picID.value()
        pixmap = QPixmap(os.path.join(path, f'frame_{"%04d" % id}{type}'))
        
        if not pixmap.isNull():
            item = scene.addPixmap(pixmap)
            self.ui.graphicsView.setScene(scene)
            # self.ui.graphicsView.fitInView(item, aspectRatioMode=1)
            self.ui.graphicsView.fitInView(item)
        else:
            print("Failed to load image.")

        # pen = QPen(QColor(0, 135, 108)) # KIT COLOR
        pen = QPen(QColor(0, 0, 255)) # BLACK
        self.line_y = 100
        self.line = scene.addLine(0, self.line_y, pixmap.width(), 100, pen)
        self.line.setZValue(1) 

        self.line_x1 = 0
        self.line_x2 = pixmap.width()

    def moveLineUp(self):
        self.line_y -= 1
        self.line.setLine(0, self.line_y, self.line.line().x2(), self.line_y)
    
    def moveLineDown(self):
        self.line_y += 1
        self.line.setLine(0, self.line_y, self.line.line().x2(), self.line_y)
    
    def moveLine10Up(self):
        self.line_y -= 10
        self.line.setLine(0, self.line_y, self.line.line().x2(), self.line_y)
    
    def moveLine10Down(self):
        self.line_y += 10
        self.line.setLine(0, self.line_y, self.line.line().x2(), self.line_y)

    def loadInput(self):
        if not self.lastFolder:
            filename, null = QFileDialog.getOpenFileName(self, filter='*.cine', options=QFileDialog.Option.ReadOnly)
            if filename: self.lastFolder = os.path.dirname(filename)
        else:
            filename, null = QFileDialog.getOpenFileName(self, directory=self.lastFolder, filter='*.cine', options=QFileDialog.Option.ReadOnly)
            if filename: self.lastFolder = os.path.dirname(filename)

        if filename.endswith('.cine'):
            print('working')
            print(filename)

            self.ui.cineLoadBar.setValue(0)
            container:av.ContainerFormat = av.open(filename)
            if not os.path.exists(os.path.join(self.path, 'global', 'currentCine')):
                os.mkdir(os.path.join(self.path, 'global', 'currentCine'))
            for item_ in os.listdir(os.path.join(self.path, 'global', 'currentCine')):
                os.remove(os.path.join(self.path, 'global', 'currentCine', item_))
            items = []
            for index, frame in enumerate(container.decode(video=0)):
                items.append((index, frame))
            
            print('Decode Done')
            container.close()
            self.lastMode = 'cine'            
            self.run_threads(items)
            self.ui.picID.setMaximum(len(items)-1)
            self.ui.picID.setValue(0)

    def loadFolder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Directory")
        if folder:
            types = ['.png', '.tif', '.tiff', '.jpeg', '.jpg']
            files = [os.path.join(folder, x) for x in os.listdir(folder) if x.endswith(types[0]) or x.endswith(types[1]) or x.endswith(types[2]) or x.endswith(types[3]) or x.endswith(types[4])]
            if len(files) > 1:
                self.currentPathFreq = folder
                self.lastMode = 'folder'
                self.ui.picID.setMaximum(len(files)-1)
                self.ui.picID.setValue(0)
            else: 
                self.currentPathFreq = None
                return
            
            self.load_image_into_graphics_view()
            
    def threadComplete(self):
        logging.info('Thread finished')
        self.ui.cineLoadBar.setValue(self.ui.cineLoadBar.value() + 1)
        self.checkButtonState()

    def freqThreadComplete(self):
        logging.info('Thread finished')
        self.ui.cineLoadBar.setValue(self.ui.cineLoadBar.value() + 1)
        if self.ui.cineLoadBar.value() >= self.ui.cineLoadBar.maximum():
            self.ui.clearRef.setEnabled(True)

    def run_threads(self, items):
        threadpool = QThreadPool.globalInstance()
        for item in items:
            worker = Worker(item, self.path, 'jpeg')
            worker.signals.finished.connect(self.threadComplete)
            threadpool.start(worker)
        
    def makeFFTList(self, value):
        self.FFTList[value[0], 1] = value[1]
        # print(self.FFTList)

    def createFreqList(self):
        if self.lastMode == None: return
        elif self.lastMode == 'cine': 
            path = os.path.join(self.path, 'global', 'currentCine')
            type = '.jpeg'
        else: 
            path = self.currentPathFreq
            type = '.png'

        files = [os.path.join(path, x) for x in os.listdir(path) if x.endswith(type)]
        if len(files) == 0: return

        threadpool = QThreadPool.globalInstance()
        self.FFTList = np.zeros((len(files), 2))
        self.FFTList[:, 0] = np.linspace(0, len(files) - 1, len(files))
        print(self.FFTList)
        i = 0
        # ref = r'M:\Duese_4\Wasser\Oben_fern_ref.tif'
        ref = self.ui.currentRef.text()
        x_start = int(self.line_x1)
        x_end = int(self.line_x2)
        y = int(self.line.line().y1())
        print(x_start, x_end, y)
        self.ui.cineLoadBar.setMaximum(len(files))
        self.ui.cineLoadBar.setValue(0)
        self.ui.clearRef.setDisabled(True)
        try:
            if ref == 'Select Reference File': raise ValueError
            refImage = cv2.imread(os.path.normpath(rf'{ref}'), cv2.IMREAD_GRAYSCALE)
        except:
            QMessageBox.information(self, 'Error', 'Please select reference image')
            return
        for file in files:
            worker = FreqWorker(i, file, ref, x_start, x_end, y, refImage)
            i += 1
            worker.signals.push.connect(self.makeFFTList)
            worker.signals.finished.connect(self.freqThreadComplete)
            threadpool.start(worker)
    
    def loadRef(self):
        if not self.lastFolder:
            filename, null = QFileDialog.getOpenFileName(self, filter='*.tif;; *.tiff;; *.png;; *.jpeg;; *.jpg', options=QFileDialog.Option.ReadOnly)
            if filename: self.lastFolder = os.path.dirname(filename)
        else:
            filename, null = QFileDialog.getOpenFileName(self, directory=self.lastFolder, filter='*.tif;; *.tiff;; *.png;; *.jpeg;; *.jpg', options=QFileDialog.Option.ReadOnly)
            if filename: self.lastFolder = os.path.dirname(filename)
        
        if not filename: return
        self.ui.currentRef.setText(filename)

    def showFFTArray(self):
        try:
            print(self.FFTList)
        except: return
        x_values = self.FFTList[:, 0]  
        y_values = self.FFTList[:, 1]  
        Fs = self.ui.frameRate.value()  
        T = 1 / Fs 
        print(np.mean(y_values))
        # t = np.arange(0, len(x_values)) * T * 1000  
        NFFT = 2**np.ceil(np.log2(len(x_values)))  
        Y = np.fft.fft(y_values, int(NFFT)) / len(x_values)
        f = Fs / 2 * np.linspace(0, 1, int(NFFT / 2) + 1)

        trace = go.Scatter(x=f[1:], y=(2*np.abs(Y[1:int(NFFT/2)+1])), mode='lines', name='FFT')
        layout = go.Layout(title=f'Single-Sided Amplitude Spectrum of y(t)\n{self.currentPathFreq}', xaxis=dict(title='Frequency (Hz)'), yaxis=dict(title='|Y(f)|'))
        fig = go.Figure(data=[trace], layout=layout)
        fig.write_html(os.path.join(self.currentPathFreq, '_freq_report.html'))
        webbrowser.open_new_tab(os.path.join(self.currentPathFreq, '_freq_report.html'))

    # Cine to Picture

    def convertFolder(self):
        file = str(QFileDialog.getExistingDirectory(self, "Select Directory"))
        self.ui.runConversion.setEnabled(True)
        self.ui.runConversion.setText('Run')
        self.ui.styleBox.setEnabled(True)
        print(file)
        # Find every .cine in folder

        items = []
        for root, dir, files in os.walk(file):
            if files:
                for item in files: 
                    if item.endswith('.cine'):
                        items.append(os.path.join(root.replace('/','\\'), item))
        print(items)
        self.ui.listWidget.clear()
        item_text = QListWidgetItem()
        item_text.setText(f'Found {len(items)} .cine files\n')
        self.ui.listWidget.addItem(item_text)

        for item in items:
            listItem = QListWidgetItem()
            listItem.setText(item.replace(r'\\', '/'))
            self.ui.listWidget.addItem(listItem)
        if len(items) > 0:
            self.cineItems = items
        else: self.cineItems = None

    def runConversion(self):
        # if 'cineItems' not in globals(): return
        if not self.cineItems: return
        self.run_threadsConversion(self.cineItems)
        
    def threadCompleteConversion(self):
        logging.info('Thread finished')
        self.ui.cineBarConversion.setValue(self.ui.cineBarConversion.value() + 1)

    def run_threadsConversion(self, items):
        print('Working on it!')
        threadpool = QThreadPool.globalInstance()
        if self.ui.keepCine.isChecked(): keep = True
        else: keep = False
        
        overwrite = self.ui.overwriteImagesCine.isChecked()

        self.ui.cineBarConversion.setMaximum(len(items)*1001)
        self.ui.cineBarConversion.setValue(0)
        style = self.ui.styleBox.currentText()
        self.ui.runConversion.setDisabled(True)
        self.ui.runConversion.setText('Running ...')
        self.ui.styleBox.setDisabled(True)

        for item in items:
            if style == '.tiff uncompressed':
                worker = WorkerConversion(item, 'tiff', keep, overwrite=overwrite)
            elif style == '.tiff compressed':
                worker = WorkerConversion(item, 'tiff', keep, True, overwrite=overwrite)
            elif style == '.png':
                worker = WorkerConversion(item, 'png', keep, overwrite=overwrite)
            elif style == '.jpeg':
                worker = WorkerConversion(item, 'jpeg', keep, overwrite=overwrite)
            else:
                worker = WorkerConversion(item, 'tiff', keep, overwrite=overwrite)
            worker.signals.finishedConversion.connect(self.threadCompleteConversion)
            threadpool.start(worker)

    # Droplet analysis
    
    def loadDropletFolder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Directory")
        if folder:
            types = ['.png', '.tif', '.tiff', '.jpeg', '.jpg']
            files = [os.path.join(folder, x) for x in os.listdir(folder) if x.endswith(types[0]) or x.endswith(types[1]) or x.endswith(types[2]) or x.endswith(types[3]) or x.endswith(types[4])]
            if len(files) > 1:
                self.currentDropletPath = folder
                self.ui.label_33.setText(folder)
                self.dropletFiles = files
            else: 
                self.currentDropletPath = None
                self.ui.label_33.setText('Make sure to select folder with images inside')

    def loadDropletRef(self):
        if not self.lastFolder:
            filename, null = QFileDialog.getOpenFileName(self, filter='*.tif;; *.tiff;; *.png;; *.jpeg;; *.jpg', options=QFileDialog.Option.ReadOnly)
            if filename: self.lastFolder = os.path.dirname(filename)
        else:
            filename, null = QFileDialog.getOpenFileName(self, directory=self.lastFolder, filter='*.tif;; *.tiff;; *.png;; *.jpeg;; *.jpg', options=QFileDialog.Option.ReadOnly)
            if filename: self.lastFolder = os.path.dirname(filename)
        
        if not filename: return
        self.dropletRefpath = filename
        self.ui.label_34.setText(filename)

    def dropletPush(self, tup:tuple):
        path, dia = tup
        self.dropletItems.append((path, dia))
        # self.ui.dropletTable.setSortingEnabled(True)
        row = self.ui.dropletTable.rowCount()
        self.ui.dropletTable.insertRow(row)
        pathItem = QTableWidgetItem(path)
        self.ui.dropletTable.setItem(row, 0, pathItem)
        diaItem = QTableWidgetItem(f'{dia}')
        diaItem.setData(1, dia)
        self.ui.dropletTable.setItem(row, 1, diaItem)
        self.sortDropletTable('d')

    def dropletFinished(self):
        self.ui.dropletProgress.setValue(self.ui.dropletProgress.value() + 1)
        if self.ui.dropletProgress.value() >= len(self.dropletFiles):
            self.ui.dropletGenerateReport.setEnabled(True)
            self.ui.dropletRun.setText('Run')
            self.ui.dropletRun.disconnect()
            self.ui.dropletRun.clicked.connect(self.dropletRun)

    def sortDropletTable(self, order='auto'):
        items = []
        for row in range(self.ui.dropletTable.rowCount()):
            items.append((self.ui.dropletTable.item(row, 0).text(), self.ui.dropletTable.item(row, 1).text()))

        if order == 'auto':
            sortedData = sorted(items, key=lambda x: float(x[1]), reverse=True)

            if sortedData[0][1] > items[-1][1]: order = 'a'
            else: order = 'd'
        
        if order == 'd': sortedData = sorted(items, key=lambda x: float(x[1]), reverse=True)
        elif order == 'a': sortedData = sorted(items, key=lambda x: float(x[1]), reverse=False)
        else: sortedData = sorted(items, key=lambda x: float(x[1]), reverse=False)
        
        # print(sortedData)
        self.ui.dropletTable.clearContents()
        self.ui.dropletTable.setRowCount(0)
        for i, (path, dia) in enumerate(sortedData):
            self.ui.dropletTable.insertRow(i)
            pathItem = QTableWidgetItem(path)
            diaItem = QTableWidgetItem(f'{dia}')
            self.ui.dropletTable.setItem(i, 0, pathItem)
            self.ui.dropletTable.setItem(i, 1, diaItem)

    def generateReport(self):
        background_img = cv2.imread(os.path.normpath(rf'{self.ui.label_34.text()}'))
        background_gray = cv2.cvtColor(background_img, cv2.COLOR_BGR2GRAY)
        export = os.path.join(self.ui.label_33.text(), 'droplet_report.pdf')
        items = []
        for i in range(10):
            items.append(self.ui.dropletTable.item(i, 0).text())

        worker = DropletsWorker(None, background_gray, self.ui.threshold.value(), self.ui.circ.value(), self.ui.dropletScale.value())
        try: worker.generateReport(items, export)
        except:
            QMessageBox.information(self, 'Error', 'Make sure to close existing Export PDF!')
            return
        os.startfile(export)

    def dropletRun(self):
        try:
            if not self.dropletRefpath: return
            if not self.currentDropletPath: return
            if not self.dropletFiles: return
        except: return

        self.dropletItems = []
              
        self.ui.dropletTable.clearContents()   
        self.ui.dropletTable.setRowCount(0)

        self.dropletthreadpool = QThreadPool.globalInstance()
        background_img = cv2.imread(os.path.normpath(rf'{self.ui.label_34.text()}'))
        background_gray = cv2.cvtColor(background_img, cv2.COLOR_BGR2GRAY)
        self.ui.dropletProgress.setMaximum(len(self.dropletFiles))
        self.ui.dropletProgress.setValue(0)
        self.ui.dropletGenerateReport.setDisabled(True)
        self.ui.dropletRun.setText('Cancel')
        # self.ui.dropletRun.setDisabled(True)
        self.ui.dropletRun.disconnect()
        self.ui.dropletRun.clicked.connect(self.stopThread)
        self.dropsAbove = []

        for i, item in enumerate(self.dropletFiles):
            worker = DropletsWorker(item, background_gray, self.ui.threshold.value(), self.ui.circ.value(), self.ui.dropletScale.value())
            worker.signals.push.connect(self.dropletPush)
            worker.signals.diasPush.connect(self.getAllDropletsAbove)
            worker.signals.finished.connect(self.dropletFinished)
            self.dropletthreadpool.start(worker)
            # print(f'Created Worker: {i}')
    
    def stopThread(self):
        self.dropletthreadpool.clear()
        self.ui.dropletRun.disconnect()
        self.ui.dropletRun.clicked.connect(self.dropletRun)
        self.ui.dropletRun.setText('Run')
    
    def getAllDropletsAbove(self, dias):
        for dia in dias:
            self.dropsAbove.append(dia)
        print(len(self.dropsAbove))

    # PDA ROUTINE
        
    def checkMatlabInstalled(self):
        try: import matlab
        except:
            self.ui.radio_mat_mode.setDisabled(True)
            self.ui.radio_mat_mode.setToolTip('Install Matlab and restart software')
            self.ui.radio_py_mode.setChecked(True)
        
    def loadPDAFolder(self, num):
        lines = [
            self.ui.PDA_Line_1,
            self.ui.PDA_Line_2,
            self.ui.PDA_Line_3,
        ]
        if self.currentPDAFolder:folder = QFileDialog.getExistingDirectory(self, "Select Directory", directory=self.currentPDAFolder)
        else: folder = QFileDialog.getExistingDirectory(self, "Select Directory")
        autoLoad = True
        validNames = [['1H', '2H', 'VP'], ['1.Halbprofil', '2.Halbprofil', 'Vollprofil']]
        if folder:
            if self.ui.actionAutomatic_Folder_Detection.isChecked():
                self.settings.set('PDA_auto_folder', True)
                sub_ = os.listdir(os.path.dirname(folder))
                top_ = os.listdir(folder)
                sub = [os.path.join(os.path.dirname(folder), x) for x in sub_ if x in validNames[0] or x in validNames[1]]
                top = [os.path.join(folder, x) for x in top_ if x in validNames[0] or x in validNames[1]]

                if len(top) > len(sub):
                    auto = top
                else: auto = sub
            else: 
                self.settings.set('PDA_auto_folder', True)
                auto = [folder]
            
            for i, a in enumerate(sorted(auto)):
                types = ['.txt']
                files = [os.path.join(a, x) for x in os.listdir(a) if x.endswith(types[0])]
                if len(files) > 1:
                    if len(auto) == 1:
                        self.currentPDAFolder = os.path.dirname(folder)
                        lines[num].setText(folder)
                    else:
                        self.currentPDAFolder = os.path.dirname(auto[0])
                        lines[i].setText(a)
                else: 
                    if len(auto) == 1:
                        lines[num].setText('Make sure to select folder with .txt files in it.')
                    else:
                        lines[i].setText('Make sure to select folder with .txt files in it.')
    
    def createItemPDA(self, ls:list):
        lines = [
            self.ui.PDA_Line_1,
            self.ui.PDA_Line_2,
            self.ui.PDA_Line_3,
        ]
        n, m, row, df_x = ls 
        item_n = QTableWidgetItem(f'{"%.3f" % n}')
        item_m =  QTableWidgetItem(f'{"%.3f" % m}')
        self.ui.PDA_table.setItem(row, 0, item_n)
        self.ui.PDA_table.setItem(row, 1, item_m)
        self.mean_m.append(m)
        self.mean_n.append(n)
        
        val1 = sum(self.mean_n)/len(self.mean_n)
        val2 = sum(self.mean_m)/len(self.mean_m)
        mean_n = QTableWidgetItem(f'{"%.3f" % val1}')
        mean_m = QTableWidgetItem(f'{"%.3f" % val2}')
        self.ui.PDA_table.setItem(3, 0, mean_n)
        self.ui.PDA_table.setItem(3, 1, mean_m)
        lines[row].setEnabled(True)
        self.dfs.append(df_x)
        self.names.append(os.path.basename(lines[row].text()))
        print(f'Basename: {os.path.basename(lines[row].text())}')
        if self.running >= 1: self.running -= 1
        if self.running == 0:
            self.ui.PDA_run.setText('Run')
            self.ui.PDA_run.setEnabled(True)
            self.ui.PDA_D32.setEnabled(True)
            self.ui.PDA_vel.setEnabled(True)
            self.ui.PDA_Vel_mean.setEnabled(True)
            self.ui.PDA_D32_mean.setEnabled(True)
            self.createTotalOut()
            try:
                self.createTotalOut()
            except:
                response = QMessageBox.question(self, 'Error', f'Please close PDA Excel file in {os.path.dirname(lines[row].text())} and press Ok afterwards', buttons=QMessageBox.StandardButton.Cancel, defaultButton=QMessageBox.StandardButton.Ok)
                if response == QMessageBox.StandardButton.Ok:
                    try: self.createTotalOut()
                    except:pass

    def runPDA(self):
        lines = [
            self.ui.PDA_Line_1,
            self.ui.PDA_Line_2,
            self.ui.PDA_Line_3,
        ]

        check = [x for x in lines if x.text() != '']
        if len(check) == 0: return

        if self.ui.PDA_cutoff.value() == 0:
            self.changeColor(self.ui.PDA_cutoff, 'red', 1000)
            return

        self.mean_n = []
        self.mean_m = []
        self.ui.PDA_table.clearContents()
        self.dfs = []
        self.m = []
        self.m_std = []
        self.n = []
        self.n_std = []
        self.names = []
        self.ui.PDA_D32.setDisabled(True)
        self.ui.PDA_vel.setDisabled(True)
        self.ui.PDA_Vel_mean.setDisabled(True)
        self.ui.PDA_D32_mean.setDisabled(True)
        
        threadpool = QThreadPool.globalInstance()
        self.ui.PDA_run.setText('Running')
        self.ui.PDA_run.setDisabled(True)
        self.running = 0
        path = MLS(os.path.join(self.path, 'global', 'scripts')).run()
        for row, line in enumerate(lines):
            if line.text().endswith('.') or line.text() == '': continue
            try:
                if self.ui.radio_mat_mode.isChecked():
                    matPath = os.path.join(self.path, 'global', f'folder{row+1}.mat')
                    worker = PDAWorker(line.text(), self.ui.PDA_cutoff.value(), self.ui.PDA_liqDens.value(), row, matPath=matPath, scriptPath=path)
                elif self.ui.radio_py_mode.isChecked():
                    worker = PDAWorker(line.text(), self.ui.PDA_cutoff.value(), self.ui.PDA_liqDens.value(), row, mode='py', scriptPath=path)
                elif self.ui.radio_py_ex_mode.isChecked():
                    worker = PDAWorker(line.text(), self.ui.PDA_cutoff.value(), self.ui.PDA_liqDens.value(), row, mode='py_ex', scriptPath=path)
                elif self.ui.radio_py_poly_mode.isChecked():
                    worker = PDAWorker(line.text(), self.ui.PDA_cutoff.value(), self.ui.PDA_liqDens.value(), row, mode='py_poly', scriptPath=path)
                worker.signals.push.connect(self.createItemPDA)
                threadpool.start(worker)
                line.setDisabled(True)
                self.running += 1
            except:
                QMessageBox.information(self, 'Error', 'Make sure to chose valid paths and if Matlab mode is selected, make sure Matlab is installed and the license is valid.')
                if self.running > 1: self.running -= 1

    def createDataFramePDA(self, mode):
        if mode == 'D32':
            d = 'D32 [µm]'
        else:
            d = 'v_z_mean [m/s]'
        
        x = 'x [mm]'
        max = 0
        max_len = 0
        
        for i, df in enumerate(self.dfs):
            if len(df) > max_len:
                max = i
        
        x_val:np.array = self.dfs[max][x].to_numpy()
        if not abs(np.max(x_val)) == abs(np.min(x_val)):
            x_val = pd.Series(data=(np.concatenate((x_val, -1*np.flip(x_val[:-1])))))

        df_push = pd.DataFrame(data=[x_val]).transpose()
        

        for i, df in enumerate(self.dfs):
            x_val:np.array = df[x].to_numpy()
            if abs(np.max(x_val)) == abs(np.min(x_val)):
                df_push = pd.concat([df_push, df[d]], ignore_index=True, axis=1)
            else: 
                d_val:np.array = df[d].to_numpy()
                df_d32_new = pd.Series(data=(np.concatenate((d_val, np.flip(d_val[:-1])))))
                df_push = pd.concat([df_push, df_d32_new], ignore_index=True, axis=1)
        
        df_push = df_push.set_index(0)
        # print(self.names)
        newCols = {col: f'{item}' for col, item in zip(df_push.columns, self.names)}
        df_push = df_push.rename(columns=newCols)

        return df_push
    
    def createD32Graph(self, mode='show'):
        df_push = self.createDataFramePDA('D32')
        df_push = df_push.reindex(sorted(df_push.columns), axis=1)  

        fig = px.line(df_push, labels={'0':'Hozizontal Position [mm]',
                                          'value':'D32 [µm]'}, markers=True)
        fig.update_layout(yaxis=dict(range=[0, df_push[self.names[0]].max()*1.1]))
        if mode == 'show' or mode == False:
            fig.show()
        else: return fig

    def createVelGraph(self, mode='show'):
        df_push = self.createDataFramePDA('Vel')
        df_push = df_push.reindex(sorted(df_push.columns), axis=1)

        fig = px.line(df_push, labels={'0':'Hozizontal Position [mm]',
                                          'value':'Axial Velocity [m/s]'}, markers=True)
        fig.update_layout(yaxis=dict(range=[0, df_push[self.names[0]].max()*1.1]))
        if mode == 'show' or mode == False:
            fig.show()
        else: return fig

    def createD32MeanGraph(self, mode:str='show'):
        df = self.createDataFramePDA('D32')
        df = df.reindex(sorted(df.columns), axis=1)
        rowMeans = df.mean(axis=1)
        rowMax = df.max(axis=1)
        rowMin = df.min(axis=1)
        df['mean'] = rowMeans
        df['pos'] = rowMax.sub(rowMeans)
        df['neg']  = rowMeans.sub(rowMin)

        fig = px.line(df['mean'], labels={'0':'Hozizontal Position [mm]',
                                          'value':'Mean D32 [µm]'}, error_y=df['pos'], error_y_minus=df['neg'], markers=True)
        fig.update_layout(yaxis=dict(range=[0, df[self.names[0]].max()*1.1]))
        if mode == 'show' or mode == False:
            fig.show()
        else: return fig

    def createVelMeanGraph(self, mode='show'):
        df = self.createDataFramePDA('Vel')
        df = df.reindex(sorted(df.columns), axis=1)
        rowMeans = df.mean(axis=1)
        rowMax = df.max(axis=1)
        rowMin = df.min(axis=1)
        df['mean'] = rowMeans
        df['pos'] = rowMax.sub(rowMeans)
        df['neg']  = rowMeans.sub(rowMin)

        fig = px.line(df['mean'], labels={'0':'Hozizontal Position [mm]',
                                          'value':'Mean Axial Velocity [m/s]'}, error_y=df['pos'], error_y_minus=df['neg'], markers=True)
        fig.update_layout(yaxis=dict(range=[0, df[self.names[0]].max()*1.1]))
        if mode == 'show' or mode == False:
            fig.show()
        else: return fig

    def createTotalOut(self):
        if not self.ui.actionGenerate_Full_Export.isChecked(): 
            self.settings.set('PDA_full_export', False)
            return
        self.settings.set('PDA_full_export', True)
        lines = [
            self.ui.PDA_Line_1,
            self.ui.PDA_Line_2,
            self.ui.PDA_Line_3,
        ]
        dirNames = []
        meanDF = pd.DataFrame(columns=['Name', 'ID_32_n', 'ID_32_m'])
        saveDF = pd.DataFrame(columns=['Name', 'ID_32_n', 'ID_32_m'])
        for x in lines:
            if x.text() == '':continue
            if os.path.dirname(x.text()) not in dirNames: dirNames.append(os.path.dirname(x.text()))
        print(dirNames)
        if len(dirNames) == 1:
            baseName = os.path.basename(dirNames[0])
            export = os.path.join(dirNames[0], f'PDA_Report_{baseName}.xlsx')
            #pd.DataFrame().to_excel(export)

            neg5 = np.arange(-100, -20, 5)
            pos2 = np.arange(-20, 20, 2)
            pos5 = np.arange(20, 105, 5)
            ind = np.concatenate((neg5, pos2, pos5))

            largest = 0
            for name, df in zip(self.names, self.dfs):
                if len(df) > largest:
                    largestName = name
                    largest = len(df)
                
            D_32_df = pd.DataFrame(columns=self.names+['mean', 'pos', 'neg', 'std'], index=ind)
            v_z_df = pd.DataFrame(columns=self.names+['mean', 'pos', 'neg', 'std'], index=ind)
            try:
                with pd.ExcelWriter(export, engine='openpyxl') as file:
                    for df, n, m, name in zip(self.dfs, self.mean_n, self.mean_m, self.names):
                        wb = file.book
                        lda = [x for x in os.listdir(os.path.join(dirNames[0], name)) if x.endswith('.lda')]
                        if lda:
                            header = lda[0]
                            offset = 3
                            maxId = len(df)+4+offset
                            startRow = 3
                        else:
                            maxId = len(df)+4
                            startRow = 0
                        
                        df.to_excel(file, sheet_name=name, index=False, startrow=startRow)
                        sheet = wb[name]
                        for index, row in df.iterrows():
                            D_32_df.loc[row['x [mm]'], name] = row['D32 [µm]']
                            v_z_df.loc[row['x [mm]'], name] = row['v_z_mean [m/s]']
                            if name != largestName:
                                D_32_df.loc[-1*row['x [mm]'], name] = row['D32 [µm]']
                                v_z_df.loc[-1*row['x [mm]'], name] = row['v_z_mean [m/s]']
                        
                        if startRow: 
                            sheet[f'A1'] = f'{os.path.join(dirNames[0], name, header)}'
                            sheet['A2'] = f'{datetime.now().strftime("%d/%m/%Y, %H:%M:%S")}'
                        sheet[f'A{maxId}'] = 'ID_32_n [µm]'
                        sheet[f'A{maxId+1}'] = 'ID_32_m [µm]'
                        sheet[f'B{maxId}'] = n
                        sheet[f'B{maxId+1}'] = m
                        wb.save(export)
                        meanDF.loc[len(meanDF)] = [name, n, m]
                        saveDF.loc[len(meanDF)] = [name, n, m]

                    print(saveDF)
                    saveDF = saveDF.set_index('Name', drop=True)
                    meanDF = meanDF.set_index('Name', drop=True)
                    meanDF = pd.concat([meanDF, pd.DataFrame(index=['mean', 'pos', 'neg', 'std'])])

                    meanDF.loc['mean'] = saveDF.mean()
                    meanDF.loc['pos'] = [saveDF['ID_32_n'].max()-meanDF.loc['mean', 'ID_32_n'], saveDF['ID_32_m'].max()-meanDF.loc['mean', 'ID_32_m']]
                    meanDF.loc['neg'] = [meanDF.loc['mean', 'ID_32_n']-saveDF['ID_32_n'].min(), meanDF.loc['mean', 'ID_32_m']-saveDF['ID_32_m'].min()]
                    meanDF.loc['std'] = saveDF.std(ddof=0)
                    
                    meanDF.transpose().to_excel(file, sheet_name='ID_32')

                    D_32_df['std'] = D_32_df.std(axis=1, ddof=0)
                    v_z_df['std'] = v_z_df.std(axis=1, ddof=0)
                    D_32_df['mean'] = D_32_df.loc[:, self.names].mean(axis=1)
                    v_z_df['mean'] = v_z_df.loc[:, self.names].mean(axis=1)
                    for index, row in D_32_df.iterrows():
                        D_32_df.loc[index, 'pos'] = row.loc[self.names].max() - row['mean']
                        D_32_df.loc[index, 'neg'] = row['mean'] - row.loc[self.names].min()
                        
                    for index, row in v_z_df.iterrows():
                        v_z_df.loc[index, 'pos'] = row.loc[self.names].max() - row['mean']
                        v_z_df.loc[index, 'neg'] = row['mean'] - row.loc[self.names].min()
                    
                    D_32_df.to_excel(file, sheet_name='D32 Export')
                    v_z_df.to_excel(file, sheet_name='v_z Export')
            except PermissionError:
                res = QMessageBox.critical(self, 'Error', f'Please close existing Excel file: {export}')
                return
            
            if self.ui.actionGenerate_and_save_Diagrams.isChecked():
                self.settings.set('PDA_diagrams', True)
                figs = []
                figs.append(self.createD32Graph(mode='save'))
                figs.append(self.createD32MeanGraph(mode='save'))
                figs.append(self.createVelGraph(mode='save'))
                figs.append(self.createVelMeanGraph(mode='save'))
                names = ['D_32', 'D_32_mean', 'vel', 'vel_mean']
                dir = os.path.join(dirNames[0], 'PDA Diagrams')

                if not os.path.exists(dir):
                    os.mkdir(dir)
                
                for fig, name in zip(figs, names):
                    fig.write_html(os.path.join(dir, f'{name}.html'))
            else: 
                self.settings.set('PDA_diagrams', False)
  
    # SPRAY ANGLE
    
    def sprayAngleInit(self):
        self.widget = MatplotlibWidget()
        self.ui.widget.setLayout(QVBoxLayout())
        self.ui.widget.layout().addWidget(self.widget)
        self.ui.angleLoadFolder.clicked.connect(self.loadSprayAngleFolder)
        self.ui.angleLoadRef.clicked.connect(self.loadSprayAngleRef)
        self.angleLastRun = None
        self.angleLastFolder = None
        self.angleLastRef = None
        self.ui.resetSprayPoints.clicked.connect(self.widget.clearPoints)
        self.setSpAState()
        self.ui.actionNormalized_Propabilty_Map.changed.connect(self.setSpAState)
        self.ui.actionMax_Width_Mode.changed.connect(self.setSpAState)
        
    def sprayAngleClear(self):
        self.widget.fig, self.widget.ax = plt.subplots()

    def angleReadFinished(self, read):
        self.probMap = mA.sumProbMap(self.probMap, read)
        self.ui.angleBar.setValue(self.ui.angleBar.value()+1)
        if self.ui.angleBar.value() == self.ui.angleBar.maximum():
            draws = [self.ui.angleMax.isChecked(), self.ui.angle10.isChecked(), self.ui.angle50.isChecked(), self.ui.angle90.isChecked()]
            if self.ui.actionNormalized_Propabilty_Map.isChecked():
                self.binaryMap, self.scaledMap = mA.createProbMap(self.probMap, np.max(self.probMap), int(self.ui.angleThreshold.value()/100*255))
            else:
                self.binaryMap, self.scaledMap = mA.createProbMap(self.probMap, self.ui.angleBar.maximum(), int(self.ui.angleThreshold.value()/100*255))
            print(np.mean(self.probMap))
            pp = mA.SprayAnglePP()
            # self.widget.clear()
            self.widget.clear()
            try:
                if self.ui.actionMax_Width_Mode.isChecked():
                    angles, image, imageRaw = pp.run(self.binaryMap, self.scaledMap, self.widget, self.ui.FLM_offset.value(), self.ui.angleTopArea.value(), draw=draws, mode='maxW', maxAngleSkip=self.ui.minSkipMaxAngle.value())
                else:
                    angles, image, imageRaw = pp.run(self.binaryMap, self.scaledMap, self.widget, self.ui.FLM_offset.value(), self.ui.angleTopArea.value(), draw=draws, mode='maxA', maxAngleSkip=self.ui.minSkipMaxAngle.value())
                self.setAngleTable(angles)
            except: pass
            self.widget.update()
            self.widget.drawPoints()

    def setAngleTable(self, angles):
        labels = [
            self.ui.angleMaxlabel,
            self.ui.angle10label,
            self.ui.angle50label,
            self.ui.angle90label
        ]

        colors = 'blue', 'green', 'yellow', 'purple'

        for label, item, color in zip(labels, angles, colors):
            label.setText(f'{"%.3f" % item} ({color})')

    def angleRun(self):
        allowedFileTypes = ['.png', '.tif', '.tiff', '.jpeg', '.jpg']
        files_ = os.listdir(self.ui.angleFolder.text())
        draws = [self.ui.angleMax.isChecked(), self.ui.angle10.isChecked(), self.ui.angle50.isChecked(), self.ui.angle90.isChecked()]

        if [self.ui.angleFolder.text(), self.ui.angleRef.text()] == self.angleLastRun:
            # self.widget.reset()
            self.widget.clear()
            if self.ui.actionNormalized_Propabilty_Map.isChecked():
                self.binaryMap, self.scaledMap = mA.createProbMap(self.probMap, np.max(self.probMap), int(self.ui.angleThreshold.value()/100*255))
            else:
                self.binaryMap, self.scaledMap = mA.createProbMap(self.probMap, self.ui.angleBar.maximum(), int(self.ui.angleThreshold.value()/100*255))
            pp = mA.SprayAnglePP()
            try:
                if self.ui.actionMax_Width_Mode.isChecked():
                    angles, image, imageRaw = pp.run(self.binaryMap, self.scaledMap, self.widget, self.ui.FLM_offset.value(), self.ui.angleTopArea.value(), draw=draws, mode='maxW', maxAngleSkip=self.ui.minSkipMaxAngle.value())
                else:
                    angles, image, imageRaw = pp.run(self.binaryMap, self.scaledMap, self.widget, self.ui.FLM_offset.value(), self.ui.angleTopArea.value(), draw=draws, mode='maxA', maxAngleSkip=self.ui.minSkipMaxAngle.value())
                self.setAngleTable(angles)
            except: pass
            self.widget.update()
            self.widget.drawPoints()
            return
        else: 
            self.angleLastRun = [self.ui.angleFolder.text(), self.ui.angleRef.text()] 
            
        files = []
        for end in allowedFileTypes:
            store = [os.path.join(self.ui.angleFolder.text(), x) for x in files_ if x.endswith(end)]
            if len(store) > len(files):
                files = store
        if len(files) == 0: return
        
        try: 
            ref = cv2.imread(os.path.normpath(rf'{self.ui.angleRef.text()}'), cv2.IMREAD_GRAYSCALE)
            self.probMap = mA.initializeProbMap(ref)
            self.angleCountMax = len(files)
            self.angleCountCurrent = 0
            self.ui.angleBar.setValue(0)
            self.ui.angleBar.setMaximum(len(files))
        except: 
            return
        
        test = cv2.imread(os.path.normpath(rf'{files[0]}'), cv2.IMREAD_GRAYSCALE)
        if np.shape(test) != np.shape(ref):
            return

        threadpool = QThreadPool.globalInstance()
        for i, item in enumerate(files):
            worker = AngleWorker(item, ref)
            worker.signals.push.connect(self.angleReadFinished)
            threadpool.start(worker)

    def loadSprayAngleFolder(self):
        if not self.angleLastFolder:
            folder = QFileDialog.getExistingDirectory(self, "Select Directory")
        else: 
            folder = QFileDialog.getExistingDirectory(self, "Select Directory", directory=self.angleLastFolder)
        if folder:
            types = ['.png', '.tif', '.tiff', '.jpeg', '.jpg']
            files = [os.path.join(folder, x) for x in os.listdir(folder) if x.endswith(types[0]) or x.endswith(types[1]) or x.endswith(types[2]) or x.endswith(types[3]) or x.endswith(types[4])]
            if len(files) > 1:
                self.angleLastFolder = folder
                self.ui.angleFolder.setText(folder)
            else:
                self.ui.angleFolder.setText('')

    def loadSprayAngleRef(self):
        if not self.angleLastRef:
            filename, null = QFileDialog.getOpenFileName(self, directory=self.angleLastRef, options=QFileDialog.Option.ReadOnly)
        else: 
            filename, null = QFileDialog.getOpenFileName(self, directory=self.path, options=QFileDialog.Option.ReadOnly)
        if not filename: return
        types = ['.png', '.tif', '.tiff', '.jpeg', '.jpg']
        match = [filename for x in types if filename.endswith(x)]
        if match:
            self.ui.angleRef.setText(match[0])
            self.lastAngleRef = os.path.dirname(match[0])
        else: 
            self.ui.angleRef.setText('')

    def setSpAState(self):
        if self.ui.actionMax_Width_Mode.isChecked():
            self.settings.set('SpA_maxW', True)
            self.ui.maxAngleDesc.setText('Angle referenced to widest point')
        else: 
            self.settings.set('SpA_maxW', False)
            self.ui.maxAngleDesc.setText('Mean Angle')
        
        if self.ui.actionNormalized_Propabilty_Map.isChecked():
            self.settings.set('SpA_norm', True)
        else: self.settings.set('SpA_norm', False)

    # Patternator

    def patternatorInit(self):
        try: self.ui.pat_chan.setValue(self.settings.get('patChan'))
        except: pass

        try: self.ui.pat_a.setValue(self.settings.get('patA'))
        except: pass

        try: self.ui.pat_b.setValue(self.settings.get('patB'))
        except: pass

        try: self.ui.pat_h.setValue(self.settings.get('patH'))
        except: pass
        
        try: self.ui.pat_op_total.setValue(self.settings.get('patOpTotal'))
        except: pass
        
        try: self.ui.pat_op.setValue(self.settings.get('patOp'))
        except: pass

        self.ui.pat_chan.valueChanged.connect(self.patternatorSaveSettings)
        self.ui.pat_a.valueChanged.connect(self.patternatorSaveSettings)
        self.ui.pat_b.valueChanged.connect(self.patternatorSaveSettings)
        self.ui.pat_h.valueChanged.connect(self.patternatorSaveSettings)
        self.ui.pat_op_total.valueChanged.connect(self.patternatorSaveSettings)
        self.ui.pat_op.valueChanged.connect(self.patternatorSaveSettings)

    def patternatorSaveSettings(self):
        try: self.settings.set('patChan', self.ui.pat_chan.value())
        except: pass

        try: self.settings.set('patA', self.ui.pat_a.value())
        except: pass

        try: self.settings.set('patB', self.ui.pat_b.value())
        except: pass

        try: self.settings.set('patH', self.ui.pat_h.value())
        except: pass

        try: self.settings.set('patOpTotal', self.ui.pat_op_total.value())
        except: pass

        try: self.settings.set('patOp', self.ui.pat_op.value())
        except: pass

    def createExcel(self):
        filename, null = QFileDialog.getSaveFileName(self, directory=os.path.join(self.path, 'patternator.xlsx'), filter='*.xlsx')
        if filename:
            try:
                pat.createExcel(filename, cols=self.ui.pat_chan.value(), tests=self.ui.pat_op_total.value(), steps=self.ui.pat_op.value())
            except:
                QMessageBox.critical(self, 'Error', 'Error during creation of the Excel template.')

    def patternatorRun(self):

        def setTooltips(sheet, tt, row=1, col=1):
            for i, t in enumerate(tt):
                sheet.cell(row, col+i).comment = Comment(t, 'DM')
        self.ui.runPat.setText('Running')
        self.ui.runPat.setDisabled(True)
        filename, null = QFileDialog.getOpenFileName(self, directory=self.path, filter='*.xlsx')
        if filename:
            try:
                patternator = pat.pat(filename, self.ui.pat_a.value(), self.ui.pat_b.value(), self.ui.pat_h.value())
                
                df = patternator.run()
                saveLoc, null = QFileDialog.getSaveFileName(self, directory=os.path.join(self.path, 'export_patternator.xlsx'), filter='*.xlsx')
                if saveLoc.endswith('.xlsx'):
                    df.to_excel(saveLoc)
                    wb = load_workbook(saveLoc)
                    ws = wb.active
                    tooltips = ['Liquid sheet thickness or nozzle name', 'Inner gas mass flow rate [kg h-1]', 'Liquid mass flow rate (important!) [kg h-1]', 
                    'Outer gas mass flow rate [kg h-1]', 'Liquid mass density (important!) [kg m-3]', 'Measuring time [s]', 'Width of PDA measurement (optional) [mm]',
                    'Width of the patternator measurement [mm]', 
                    'Relative mass loss towards the outside of the patternator measurement width [%]',
                    'Absolute mass loss towards the outside of the patternator measurement width [kg h-1]',
                    'Relative mass loss towards the outside of the patternator measurement width using the Lorentz-fitted fluxes [%]',
                    'Relative mass loss towards the outside of the PDA measurement width using the Lorentz-fitted fluxes [%]',
                    'R^2 Value of the Lorentz-Fit']

                    setTooltips(ws, tooltips, 1, 2)
                    wb.save(saveLoc)
                else: raise PermissionError
            except: 
                QMessageBox.critical(self, 'Error', 'Error during creation handling of the data.')
            
            self.ui.runPat.setText('Select Measurement Data')
            self.ui.runPat.setEnabled(True)

def show_error_popup():
    # app = QApplication([])
    error_popup = QMessageBox()
    error_popup.setIcon(QMessageBox.Icon.Critical)
    error_popup.setWindowTitle("Error")
    error_popup.setText(f"An error occurred!\nA crash log can be found in {os.path.join(os.path.expanduser('~'), 'Atomizer Toolbox', 'logs')}")
    error_popup.setStandardButtons(QMessageBox.StandardButton.Ok)
    error_popup.exec()
    app.quit()        
    os._exit(1)        

def excepthook(exc_type, exc_value, traceback):
    # Log unhandled exceptions
    print('ERROR')
    logging.exception("Unhandled exception", exc_info=(exc_type, exc_value, traceback))
    print('Logging Done')
    os.startfile(os.path.join(os.path.expanduser('~'), 'Atomizer Toolbox', 'logs'))
    os._exit(1)  

if __name__ == '__main__':
    try: os.mkdir(os.path.join(os.path.expanduser('~'), 'Atomizer Toolbox', 'logs'))
    except: 
        try: 
            os.mkdir(os.path.join(os.path.expanduser('~'), 'Atomizer Toolbox'))
            os.mkdir(os.path.join(os.path.expanduser('~'), 'Atomizer Toolbox', 'logs'))
        except: pass

    log_directory = os.path.join(os.path.expanduser('~'), 'Atomizer Toolbox', 'logs')
    log_file = f"{log_directory}/crash.log"

    logging.basicConfig(filename=log_file, level=logging.ERROR,
                    format='%(asctime)s - %(levelname)s - %(message)s')


    sys.excepthook = excepthook

    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = UI()
    window.show()
    sys.exit(app.exec())
